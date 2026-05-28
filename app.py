"""
SACCOFinance LMS - Flask REST API
All routes: auth, members, loans, savings, repayments, reports, notifications, audit
"""
from flask import Flask, request, jsonify, g, send_from_directory, make_response, send_file
import os, json, sqlite3, base64, re
from io import BytesIO
from datetime import datetime, date
import time as pytime
from html import escape

from database import get_db, init_db, hash_password as db_hash
from auth     import (hash_password, generate_token, decode_token,
                      login_required, roles_required, gen_id)
from calculator import build_schedule, calculate_penalty, loan_summary

# ── App setup ────────────────────────────────────────────────────────────────
FRONTEND_DIR = os.path.dirname(__file__)
REACT_DIST_DIR = os.path.join(FRONTEND_DIR, "frontend", "dist")
app = Flask(__name__, static_folder=FRONTEND_DIR, static_url_path="")

# Manual CORS (no flask-cors available)
@app.after_request
def add_cors(response):
    response.headers["Access-Control-Allow-Origin"]  = "*"
    response.headers["Access-Control-Allow-Headers"] = "Content-Type, Authorization"
    response.headers["Access-Control-Allow-Methods"] = "GET, POST, PUT, PATCH, DELETE, OPTIONS"
    return response

@app.before_request
def handle_options():
    if request.method == "OPTIONS":
        from flask import make_response
        return make_response("", 204)

# ── Helpers ──────────────────────────────────────────────────────────────────
def row_to_dict(row) -> dict:
    return dict(row) if row else {}

def rows_to_list(rows) -> list:
    return [dict(r) for r in rows]

def audit(action: str, module: str, details: str = ""):
    try:
        user = g.get("user", {})
        db   = get_db()
        db.execute(
            "INSERT INTO audit_logs (user_id,user_name,action,module,details,ip_address) VALUES (?,?,?,?,?,?)",
            (user.get("sub"), user.get("name","System"), action, module, details,
             request.remote_addr or "unknown"),
        )
        db.commit(); db.close()
    except Exception:
        pass

def success(data=None, msg="OK", code=200):
    return jsonify({"success": True,  "message": msg, "data": data}), code

def error(msg="Error", code=400):
    return jsonify({"success": False, "error": msg}), code


PASSWORD_POLICY_MESSAGE = (
    "Password must be at least 8 characters and include uppercase, lowercase, number, and special character."
)
VALID_USER_ROLES = {"admin", "officer", "accountant", "cashier"}


def validate_password_strength(password: str) -> str | None:
    pwd = str(password or "")
    if len(pwd) < 8:
        return PASSWORD_POLICY_MESSAGE
    if not re.search(r"[A-Z]", pwd):
        return PASSWORD_POLICY_MESSAGE
    if not re.search(r"[a-z]", pwd):
        return PASSWORD_POLICY_MESSAGE
    if not re.search(r"\d", pwd):
        return PASSWORD_POLICY_MESSAGE
    if not re.search(r"[^A-Za-z0-9]", pwd):
        return PASSWORD_POLICY_MESSAGE
    return None


def validate_user_role(role: str) -> str | None:
    if (role or "").strip().lower() not in VALID_USER_ROLES:
        return "Role must be one of: admin, officer, accountant, cashier"
    return None


def _int_env(name: str, default: int, minimum: int) -> int:
    try:
        return max(minimum, int(os.environ.get(name, str(default))))
    except (TypeError, ValueError):
        return max(minimum, default)


LOGIN_MAX_ATTEMPTS = _int_env("LOGIN_MAX_ATTEMPTS", 5, 1)
LOGIN_WINDOW_SECONDS = _int_env("LOGIN_WINDOW_SECONDS", 600, 60)
LOGIN_BLOCK_SECONDS = _int_env("LOGIN_BLOCK_SECONDS", 900, 60)
_LOGIN_ATTEMPTS: dict[str, dict] = {}


def _cleanup_login_attempts(now_ts: float) -> None:
    expired = []
    for key, rec in _LOGIN_ATTEMPTS.items():
        blocked_until = float(rec.get("blocked_until", 0) or 0)
        last_seen = float(rec.get("last_seen", 0) or 0)
        if blocked_until and blocked_until > now_ts:
            continue
        if now_ts - last_seen > LOGIN_WINDOW_SECONDS * 2:
            expired.append(key)
    for key in expired:
        _LOGIN_ATTEMPTS.pop(key, None)


def _login_attempt_key(username: str) -> str:
    xff = (request.headers.get("X-Forwarded-For") or "").strip()
    ip = (xff.split(",")[0].strip() if xff else "") or (request.remote_addr or "unknown")
    return f"{ip}:{username.lower().strip()}"


def _login_block_remaining_seconds(attempt_key: str) -> int:
    now_ts = pytime.time()
    rec = _LOGIN_ATTEMPTS.get(attempt_key)
    if not rec:
        return 0
    blocked_until = float(rec.get("blocked_until", 0) or 0)
    if blocked_until <= now_ts:
        return 0
    return int(blocked_until - now_ts)


def _record_failed_login(attempt_key: str) -> int:
    now_ts = pytime.time()
    rec = _LOGIN_ATTEMPTS.get(attempt_key, {"attempts": 0, "window_start": now_ts, "blocked_until": 0, "last_seen": now_ts})
    window_start = float(rec.get("window_start", now_ts) or now_ts)
    if now_ts - window_start > LOGIN_WINDOW_SECONDS:
        rec["attempts"] = 0
        rec["window_start"] = now_ts
        rec["blocked_until"] = 0
    rec["attempts"] = int(rec.get("attempts", 0) or 0) + 1
    rec["last_seen"] = now_ts
    if rec["attempts"] >= LOGIN_MAX_ATTEMPTS:
        rec["blocked_until"] = now_ts + LOGIN_BLOCK_SECONDS
    _LOGIN_ATTEMPTS[attempt_key] = rec
    _cleanup_login_attempts(now_ts)
    return _login_block_remaining_seconds(attempt_key)


def _clear_login_attempts(attempt_key: str) -> None:
    _LOGIN_ATTEMPTS.pop(attempt_key, None)

def get_settings_dict():
    db = get_db()
    rows = rows_to_list(db.execute("SELECT key,value FROM app_settings").fetchall())
    db.close()
    settings = {
        "sacco_name": "SACCOFinance",
        "logo_text": "SF",
        "logo_image": "",
        "logo_url": "",
        "address": "",
        "phone": "",
        "account_opening_balance": "0",
    }
    settings.update({r["key"]: r["value"] for r in rows})
    return settings

def get_account_opening_balance(db) -> float:
    opening_balance_row = db.execute(
        "SELECT value FROM app_settings WHERE key='account_opening_balance'"
    ).fetchone()
    try:
        return float((opening_balance_row["value"] if opening_balance_row else 0) or 0)
    except (TypeError, ValueError):
        return 0.0

def set_account_opening_balance(db, value: float) -> float:
    new_value = float(value or 0)
    db.execute(
        "INSERT INTO app_settings (key,value,updated_at) VALUES (?,?,datetime('now')) "
        "ON CONFLICT(key) DO UPDATE SET value=excluded.value, updated_at=datetime('now')",
        ("account_opening_balance", str(new_value)),
    )
    return new_value

def adjust_account_opening_balance(db, delta: float) -> float:
    current = get_account_opening_balance(db)
    return set_account_opening_balance(db, current + float(delta or 0))

def build_monthly_account_report(db) -> dict:
    monthly_raw = rows_to_list(db.execute(
        """
        SELECT month,
               COALESCE(SUM(savings_collections),0) AS savings_collections,
               COALESCE(SUM(loan_repayments),0) AS loan_repayments,
               COALESCE(SUM(loan_disbursed),0) AS loan_disbursed,
               COALESCE(SUM(expenses),0) AS expenses
        FROM (
            SELECT strftime('%Y-%m', txn_date) AS month,
                   amount AS savings_collections,
                   0 AS loan_repayments,
                   0 AS loan_disbursed,
                   0 AS expenses
            FROM savings_transactions
            WHERE type='deposit'

            UNION ALL

            SELECT strftime('%Y-%m', payment_date) AS month,
                   0 AS savings_collections,
                   amount AS loan_repayments,
                   0 AS loan_disbursed,
                   0 AS expenses
            FROM repayments

            UNION ALL

            SELECT strftime('%Y-%m', disbursed_date) AS month,
                   0 AS savings_collections,
                   0 AS loan_repayments,
                   amount AS loan_disbursed,
                   0 AS expenses
            FROM loans
            WHERE disbursed_date IS NOT NULL

            UNION ALL

            SELECT strftime('%Y-%m', expense_date) AS month,
                   0 AS savings_collections,
                   0 AS loan_repayments,
                   0 AS loan_disbursed,
                   amount AS expenses
            FROM expense_transactions
        )
        WHERE month IS NOT NULL AND month != ''
        GROUP BY month
        ORDER BY month ASC
        """
    ).fetchall())

    current_balance = get_account_opening_balance(db)
    if not monthly_raw:
        return {
            "opening_balance": current_balance,
            "closing_balance": current_balance,
            "months": [],
            "totals": {
                "savings_collections": 0.0,
                "loan_repayments": 0.0,
                "inflow": 0.0,
                "loan_disbursed": 0.0,
                "expenses": 0.0,
                "outflow": 0.0,
                "net": 0.0,
            },
        }

    by_month = {r["month"]: r for r in monthly_raw if r.get("month")}
    overall_net = 0.0
    for row in by_month.values():
        savings = float(row.get("savings_collections") or 0)
        repaid = float(row.get("loan_repayments") or 0)
        disbursed = float(row.get("loan_disbursed") or 0)
        expenses = float(row.get("expenses") or 0)
        overall_net += (savings + repaid) - (disbursed + expenses)

    opening_balance = current_balance - overall_net
    month_keys = sorted(by_month.keys())
    start_month = datetime.strptime(f"{month_keys[0]}-01", "%Y-%m-%d").date()
    end_month = datetime.strptime(f"{month_keys[-1]}-01", "%Y-%m-%d").date()

    def next_month(d: date) -> date:
        if d.month == 12:
            return date(d.year + 1, 1, 1)
        return date(d.year, d.month + 1, 1)

    rows = []
    running_balance = opening_balance
    totals = {
        "savings_collections": 0.0,
        "loan_repayments": 0.0,
        "inflow": 0.0,
        "loan_disbursed": 0.0,
        "expenses": 0.0,
        "outflow": 0.0,
        "net": 0.0,
    }
    cursor = start_month
    while cursor <= end_month:
        month_key = cursor.strftime("%Y-%m")
        source = by_month.get(month_key, {})
        savings_collections = float(source.get("savings_collections") or 0)
        loan_repayments = float(source.get("loan_repayments") or 0)
        loan_disbursed = float(source.get("loan_disbursed") or 0)
        expenses = float(source.get("expenses") or 0)
        inflow = savings_collections + loan_repayments
        outflow = loan_disbursed + expenses
        net = inflow - outflow
        opening = running_balance
        closing = opening + net

        rows.append({
            "month": month_key,
            "opening_balance": opening,
            "savings_collections": savings_collections,
            "loan_repayments": loan_repayments,
            "inflow": inflow,
            "loan_disbursed": loan_disbursed,
            "expenses": expenses,
            "outflow": outflow,
            "net": net,
            "closing_balance": closing,
        })

        totals["savings_collections"] += savings_collections
        totals["loan_repayments"] += loan_repayments
        totals["inflow"] += inflow
        totals["loan_disbursed"] += loan_disbursed
        totals["expenses"] += expenses
        totals["outflow"] += outflow
        totals["net"] += net

        running_balance = closing
        cursor = next_month(cursor)

    return {
        "opening_balance": opening_balance,
        "closing_balance": running_balance,
        "months": rows,
        "totals": totals,
    }


ACCOUNTING_ACCOUNTS = [
    {"code": "1000", "name": "Main Cash / Bank", "type": "asset"},
    {"code": "1100", "name": "Loan Portfolio Receivable", "type": "asset"},
    {"code": "2100", "name": "Member Savings Liability", "type": "liability"},
    {"code": "3000", "name": "Opening Capital / Equity", "type": "equity"},
    {"code": "4000", "name": "Interest Income", "type": "income"},
    {"code": "5100", "name": "Operating Expenses", "type": "expense"},
]
ACCOUNT_BY_CODE = {a["code"]: a for a in ACCOUNTING_ACCOUNTS}


def _journal_line(account_code: str, debit: float = 0, credit: float = 0) -> dict:
    account = ACCOUNT_BY_CODE[account_code]
    return {
        "account_code": account_code,
        "account_name": account["name"],
        "account_type": account["type"],
        "debit": round(float(debit or 0), 2),
        "credit": round(float(credit or 0), 2),
    }


def _add_journal_entry(entries: list, source: str, source_id: str, entry_date: str, description: str, lines: list) -> None:
    clean_lines = [line for line in lines if float(line.get("debit") or 0) or float(line.get("credit") or 0)]
    if not clean_lines:
        return
    total_debit = round(sum(float(line["debit"] or 0) for line in clean_lines), 2)
    total_credit = round(sum(float(line["credit"] or 0) for line in clean_lines), 2)
    entries.append({
        "id": f"JE{len(entries) + 1:06d}",
        "source": source,
        "source_id": source_id,
        "entry_date": clean_date(entry_date),
        "description": description,
        "total_debit": total_debit,
        "total_credit": total_credit,
        "balanced": abs(total_debit - total_credit) < 0.01,
        "lines": clean_lines,
    })


def _repayment_allocations(db) -> dict:
    schedules = {}
    for row in rows_to_list(db.execute(
        "SELECT loan_id, installment, principal, interest FROM loan_schedule ORDER BY loan_id, installment"
    ).fetchall()):
        schedules.setdefault(row["loan_id"], []).append({
            "principal_remaining": float(row.get("principal") or 0),
            "interest_remaining": float(row.get("interest") or 0),
        })

    allocations = {}
    repayments = rows_to_list(db.execute(
        "SELECT id, loan_id, amount FROM repayments ORDER BY loan_id, payment_date, created_at"
    ).fetchall())
    for repayment in repayments:
        remaining = float(repayment.get("amount") or 0)
        principal = 0.0
        interest = 0.0
        for slot in schedules.get(repayment["loan_id"], []):
            if remaining <= 0:
                break
            pay_interest = min(remaining, slot["interest_remaining"])
            slot["interest_remaining"] -= pay_interest
            interest += pay_interest
            remaining -= pay_interest
            if remaining <= 0:
                break
            pay_principal = min(remaining, slot["principal_remaining"])
            slot["principal_remaining"] -= pay_principal
            principal += pay_principal
            remaining -= pay_principal
        if remaining > 0:
            principal += remaining
        allocations[repayment["id"]] = {
            "principal": round(principal, 2),
            "interest": round(interest, 2),
        }
    return allocations


def build_accounting_journal(db) -> list:
    entries = []

    for row in rows_to_list(db.execute(
        """SELECT st.*, m.name AS member_name
           FROM savings_transactions st
           JOIN members m ON m.id=st.member_id
           ORDER BY st.txn_date, st.created_at"""
    ).fetchall()):
        amount = float(row.get("amount") or 0)
        if row.get("type") == "withdrawal":
            _add_journal_entry(entries, "savings_withdrawal", row["id"], row["txn_date"],
                f"Savings withdrawal - {row.get('member_name') or row['member_id']}",
                [_journal_line("2100", debit=amount), _journal_line("1000", credit=amount)])
        else:
            _add_journal_entry(entries, "savings_deposit", row["id"], row["txn_date"],
                f"Savings deposit - {row.get('member_name') or row['member_id']}",
                [_journal_line("1000", debit=amount), _journal_line("2100", credit=amount)])

    for row in rows_to_list(db.execute(
        """SELECT l.*, m.name AS member_name
           FROM loans l JOIN members m ON m.id=l.member_id
           WHERE l.disbursed_date IS NOT NULL
           ORDER BY l.disbursed_date, l.created_at"""
    ).fetchall()):
        amount = float(row.get("amount") or 0)
        _add_journal_entry(entries, "loan_disbursement", row["id"], row["disbursed_date"],
            f"Loan disbursement - {row.get('member_name') or row['member_id']}",
            [_journal_line("1100", debit=amount), _journal_line("1000", credit=amount)])

    allocations = _repayment_allocations(db)
    for row in rows_to_list(db.execute(
        """SELECT r.*, m.name AS member_name
           FROM repayments r JOIN members m ON m.id=r.member_id
           ORDER BY r.payment_date, r.created_at"""
    ).fetchall()):
        amount = float(row.get("amount") or 0)
        allocation = allocations.get(row["id"], {"principal": amount, "interest": 0})
        principal = min(amount, float(allocation.get("principal") or 0))
        interest = max(0.0, amount - principal)
        _add_journal_entry(entries, "loan_repayment", row["id"], row["payment_date"],
            f"Loan repayment {row['loan_id']} - {row.get('member_name') or row['member_id']}",
            [
                _journal_line("1000", debit=amount),
                _journal_line("1100", credit=principal),
                _journal_line("4000", credit=interest),
            ])

    for row in rows_to_list(db.execute(
        """SELECT et.*, ea.name AS account_name
           FROM expense_transactions et
           JOIN expense_accounts ea ON ea.id=et.account_id
           ORDER BY et.expense_date, et.created_at"""
    ).fetchall()):
        amount = float(row.get("amount") or 0)
        _add_journal_entry(entries, "expense", row["id"], row["expense_date"],
            f"Expense - {row.get('account_name') or 'Operating expense'}",
            [_journal_line("5100", debit=amount), _journal_line("1000", credit=amount)])

    cash_delta = 0.0
    for entry in entries:
        for line in entry["lines"]:
            if line["account_code"] == "1000":
                cash_delta += float(line["debit"] or 0) - float(line["credit"] or 0)
    current_cash = get_account_opening_balance(db)
    opening_cash = round(current_cash - cash_delta, 2)
    if abs(opening_cash) >= 0.01:
        if opening_cash > 0:
            lines = [_journal_line("1000", debit=opening_cash), _journal_line("3000", credit=opening_cash)]
        else:
            lines = [_journal_line("3000", debit=abs(opening_cash)), _journal_line("1000", credit=abs(opening_cash))]
        _add_journal_entry(entries, "opening_balance", "OPENING", "1900-01-01",
            "Opening balance brought forward", lines)

    entries.sort(key=lambda e: (e["entry_date"], e["source"], e["source_id"]))
    for idx, entry in enumerate(entries, start=1):
        entry["id"] = f"JE{idx:06d}"
    return entries


def _filter_journal(entries: list, date_from: str = "", date_to: str = "", source: str = "") -> list:
    if date_from:
        date_from = clean_date(date_from)
        entries = [entry for entry in entries if entry["entry_date"] >= date_from]
    if date_to:
        date_to = clean_date(date_to)
        entries = [entry for entry in entries if entry["entry_date"] <= date_to]
    if source:
        entries = [entry for entry in entries if entry["source"] == source]
    return entries


def _account_balances(entries: list) -> dict:
    balances = {
        account["code"]: {
            **account,
            "debit": 0.0,
            "credit": 0.0,
            "net": 0.0,
            "balance": 0.0,
        }
        for account in ACCOUNTING_ACCOUNTS
    }
    for entry in entries:
        for line in entry["lines"]:
            row = balances[line["account_code"]]
            row["debit"] += float(line.get("debit") or 0)
            row["credit"] += float(line.get("credit") or 0)
    for row in balances.values():
        row["debit"] = round(row["debit"], 2)
        row["credit"] = round(row["credit"], 2)
        row["net"] = round(row["debit"] - row["credit"], 2)
        if row["type"] in ("asset", "expense"):
            row["balance"] = row["net"]
        else:
            row["balance"] = round(row["credit"] - row["debit"], 2)
    return balances


def build_trial_balance(db, date_from: str = "", date_to: str = "") -> dict:
    entries = _filter_journal(build_accounting_journal(db), date_from, date_to)
    rows = []
    total_debit = 0.0
    total_credit = 0.0
    for row in _account_balances(entries).values():
        net = float(row["net"] or 0)
        debit_balance = net if net > 0 else 0.0
        credit_balance = abs(net) if net < 0 else 0.0
        total_debit += debit_balance
        total_credit += credit_balance
        rows.append({**row, "debit_balance": round(debit_balance, 2), "credit_balance": round(credit_balance, 2)})
    return {
        "rows": rows,
        "totals": {
            "debit": round(total_debit, 2),
            "credit": round(total_credit, 2),
            "balanced": abs(total_debit - total_credit) < 0.01,
        },
    }


def build_profit_loss(db, date_from: str = "", date_to: str = "") -> dict:
    balances = _account_balances(_filter_journal(build_accounting_journal(db), date_from, date_to))
    income_rows = [row for row in balances.values() if row["type"] == "income" and abs(row["balance"]) >= 0.01]
    expense_rows = [row for row in balances.values() if row["type"] == "expense" and abs(row["balance"]) >= 0.01]
    income = sum(float(row["balance"] or 0) for row in income_rows)
    expenses = sum(float(row["balance"] or 0) for row in expense_rows)
    return {
        "income": income_rows,
        "expenses": expense_rows,
        "totals": {
            "income": round(income, 2),
            "expenses": round(expenses, 2),
            "net_profit": round(income - expenses, 2),
        },
    }


def build_balance_sheet(db, date_to: str = "") -> dict:
    entries = _filter_journal(build_accounting_journal(db), "", date_to)
    balances = _account_balances(entries)
    assets = [row for row in balances.values() if row["type"] == "asset" and abs(row["balance"]) >= 0.01]
    liabilities = [row for row in balances.values() if row["type"] == "liability" and abs(row["balance"]) >= 0.01]
    equity = [row for row in balances.values() if row["type"] == "equity" and abs(row["balance"]) >= 0.01]
    pnl = build_profit_loss(db, "", date_to)
    retained = pnl["totals"]["net_profit"]
    if abs(retained) >= 0.01:
        equity.append({
            "code": "3900",
            "name": "Retained Earnings / Current Profit",
            "type": "equity",
            "debit": 0.0,
            "credit": 0.0,
            "net": -retained,
            "balance": retained,
        })
    total_assets = sum(float(row["balance"] or 0) for row in assets)
    total_liabilities = sum(float(row["balance"] or 0) for row in liabilities)
    total_equity = sum(float(row["balance"] or 0) for row in equity)
    return {
        "assets": assets,
        "liabilities": liabilities,
        "equity": equity,
        "totals": {
            "assets": round(total_assets, 2),
            "liabilities": round(total_liabilities, 2),
            "equity": round(total_equity, 2),
            "liabilities_and_equity": round(total_liabilities + total_equity, 2),
            "balanced": abs(total_assets - (total_liabilities + total_equity)) < 0.01,
        },
    }


def build_cash_flow_statement(db, date_from: str = "", date_to: str = "") -> dict:
    entries = _filter_journal(build_accounting_journal(db), date_from, date_to)
    by_month = {}
    for entry in entries:
        if entry["source"] == "opening_balance":
            continue
        month = entry["entry_date"][:7]
        row = by_month.setdefault(month, {
            "month": month,
            "operating": 0.0,
            "investing": 0.0,
            "financing": 0.0,
            "net_cash_flow": 0.0,
        })
        cash_line = next((line for line in entry["lines"] if line["account_code"] == "1000"), None)
        cash_delta = float(cash_line.get("debit") or 0) - float(cash_line.get("credit") or 0) if cash_line else 0.0
        if entry["source"] == "loan_repayment":
            interest = sum(float(line.get("credit") or 0) for line in entry["lines"] if line["account_code"] == "4000")
            principal = cash_delta - interest
            row["operating"] += interest
            row["investing"] += principal
        elif entry["source"] == "expense":
            row["operating"] += cash_delta
        elif entry["source"] == "loan_disbursement":
            row["investing"] += cash_delta
        elif entry["source"] in ("savings_deposit", "savings_withdrawal"):
            row["financing"] += cash_delta
        row["net_cash_flow"] += cash_delta

    rows = []
    totals = {"operating": 0.0, "investing": 0.0, "financing": 0.0, "net_cash_flow": 0.0}
    for row in [by_month[key] for key in sorted(by_month.keys())]:
        for key in totals:
            row[key] = round(row[key], 2)
            totals[key] += row[key]
        rows.append(row)
    return {"months": rows, "totals": {k: round(v, 2) for k, v in totals.items()}}

def pdf_escape(value) -> str:
    return str(value).replace("\\", "\\\\").replace("(", "\\(").replace(")", "\\)")

def pdf_clean(value) -> str:
    return str(value if value is not None else "").encode("latin-1", "replace").decode("latin-1")

def pdf_money(value) -> str:
    return f"KES {float(value or 0):,.2f}"

def pdf_date(value) -> str:
    raw = str(value or "").strip()
    if not raw:
        return "-"
    try:
        return datetime.fromisoformat(raw[:10]).strftime("%d %b %Y")
    except Exception:
        return raw[:10]

def pdf_text(content, x, y, size=9, font="F1", color=(24, 32, 51), align="left", max_chars=None):
    text = pdf_clean(content)
    if max_chars and len(text) > max_chars:
        text = text[: max_chars - 1].rstrip() + "..."
    width = len(text) * size * 0.48
    if align == "right":
        x = x - width
    elif align == "center":
        x = x - (width / 2)
    r, g, b = [c / 255 for c in color]
    return f"BT /{font} {size} Tf {r:.3f} {g:.3f} {b:.3f} rg {x:.2f} {y:.2f} Td ({pdf_escape(text)}) Tj ET"

def pdf_rect(x, y, w, h, fill=None, stroke=None, line_width=0.6):
    ops = ["q"]
    if fill:
        r, g, b = [c / 255 for c in fill]
        ops.append(f"{r:.3f} {g:.3f} {b:.3f} rg {x:.2f} {y:.2f} {w:.2f} {h:.2f} re f")
    if stroke:
        r, g, b = [c / 255 for c in stroke]
        ops.append(f"{line_width:.2f} w {r:.3f} {g:.3f} {b:.3f} RG {x:.2f} {y:.2f} {w:.2f} {h:.2f} re S")
    ops.append("Q")
    return "\n".join(ops)

def pdf_line(x1, y1, x2, y2, color=(216, 222, 232), line_width=0.6):
    r, g, b = [c / 255 for c in color]
    return f"q {line_width:.2f} w {r:.3f} {g:.3f} {b:.3f} RG {x1:.2f} {y1:.2f} m {x2:.2f} {y2:.2f} l S Q"

def pdf_jpeg_from_data_url(value):
    match = re.match(r"^data:image/jpeg;base64,(.+)$", str(value or ""), re.I)
    if not match:
        return None
    try:
        return base64.b64decode(match.group(1), validate=True)
    except Exception:
        return None

def make_pdf_from_pages(page_streams, title="Loan Statement", logo_image="") -> bytes:
    page_count = len(page_streams)
    logo_bytes = pdf_jpeg_from_data_url(logo_image)
    font_start = 3 + page_count
    image_start = font_start + 2
    content_start = image_start + (1 if logo_bytes else 0)
    page_refs = " ".join(f"{3 + i} 0 R" for i in range(page_count))
    objects = [
        b"<< /Type /Catalog /Pages 2 0 R >>",
        f"<< /Type /Pages /Kids [{page_refs}] /Count {page_count} >>".encode("ascii"),
    ]
    for i in range(page_count):
        objects.append(
            f"<< /Type /Page /Parent 2 0 R /MediaBox [0 0 595 842] "
            f"/Resources << /Font << /F1 {font_start} 0 R /F2 {font_start + 1} 0 R >> "
            f"{('/XObject << /Logo ' + str(image_start) + ' 0 R >>') if logo_bytes else ''} >> "
            f"/Contents {content_start + i} 0 R >>"
        .encode("ascii"))
    objects.extend([
        b"<< /Type /Font /Subtype /Type1 /BaseFont /Helvetica >>",
        b"<< /Type /Font /Subtype /Type1 /BaseFont /Helvetica-Bold >>",
    ])
    if logo_bytes:
        objects.append(
            b"<< /Type /XObject /Subtype /Image /Width 512 /Height 512 "
            b"/ColorSpace /DeviceRGB /BitsPerComponent 8 /Filter /DCTDecode /Length "
            + str(len(logo_bytes)).encode("ascii") + b" >>\nstream\n" + logo_bytes + b"\nendstream"
        )
    for stream_text in page_streams:
        stream = stream_text.encode("latin-1", "replace")
        objects.append(b"<< /Length " + str(len(stream)).encode("ascii") + b" >>\nstream\n" + stream + b"\nendstream")

    pdf = BytesIO()
    pdf.write(b"%PDF-1.4\n")
    offsets = []
    for i, obj in enumerate(objects, 1):
        offsets.append(pdf.tell())
        pdf.write(f"{i} 0 obj\n".encode("ascii"))
        pdf.write(obj)
        pdf.write(b"\nendobj\n")
    xref = pdf.tell()
    pdf.write(f"xref\n0 {len(objects) + 1}\n".encode("ascii"))
    pdf.write(b"0000000000 65535 f \n")
    for offset in offsets:
        pdf.write(f"{offset:010d} 00000 n \n".encode("ascii"))
    pdf.write(f"trailer << /Size {len(objects) + 1} /Root 1 0 R >>\nstartxref\n{xref}\n%%EOF".encode("ascii"))
    return pdf.getvalue()

def make_simple_pdf(lines, title="Loan Statement") -> bytes:
    y = 790
    ops = [pdf_text(title, 50, y, 14, "F2")]
    y -= 22
    for line in lines:
        ops.append(pdf_text(line, 50, y, 10))
        y -= 14
    return make_pdf_from_pages(["\n".join(ops)], title)

def get_loan_statement_data(loan_id):
    db = get_db()
    loan = row_to_dict(db.execute(
        """SELECT l.*, m.name as member_name, m.phone as member_phone, m.address as member_address, m.member_type
           FROM loans l JOIN members m ON l.member_id=m.id WHERE l.id=?""", (loan_id,)
    ).fetchone())
    if not loan:
        db.close()
        return None
    schedule = rows_to_list(db.execute("SELECT * FROM loan_schedule WHERE loan_id=? ORDER BY installment", (loan_id,)).fetchall())
    repays = rows_to_list(db.execute("SELECT * FROM repayments WHERE loan_id=? ORDER BY payment_date", (loan_id,)).fetchall())
    db.close()
    return {
        "loan": loan,
        "schedule": schedule,
        "repays": repays,
        "settings": get_settings_dict(),
        "summary": loan_summary(loan, schedule),
    }

def build_statement_pdf(loan_id, data) -> bytes:
    loan = data["loan"]
    summary = data["summary"]
    settings = data["settings"]
    brand = settings.get("sacco_name") or "SACCOFinance"
    pages = []
    ops = []
    margin = 42
    width = 511
    blue = (23, 32, 51)
    muted = (100, 116, 139)
    border = (214, 221, 232)
    soft = (247, 249, 252)
    accent = (37, 99, 235)

    def add_header(page_no=1):
        o = [
            pdf_rect(0, 782, 595, 60, fill=blue),
            pdf_rect(margin, 798, 34, 34, fill=accent),
            ("q 34 0 0 34 %.2f %.2f cm /Logo Do Q" % (margin, 798)) if pdf_jpeg_from_data_url(settings.get("logo_image")) else pdf_text((settings.get("logo_text") or brand[:2]).upper()[:3], margin + 17, 810, 11, "F2", (255, 255, 255), "center"),
            pdf_text(brand, 86, 818, 16, "F2", (255, 255, 255), max_chars=34),
            pdf_text(settings.get("address") or "Loan management statement", 86, 803, 8, "F1", (213, 226, 255), max_chars=58),
            pdf_text(settings.get("phone") or "", 86, 791, 8, "F1", (213, 226, 255), max_chars=42),
            pdf_text("LOAN STATEMENT", 553, 818, 12, "F2", (255, 255, 255), "right"),
            pdf_text(f"Statement No: {loan_id}", 553, 803, 8, "F1", (213, 226, 255), "right"),
            pdf_text(f"Issued: {pdf_date(date.today().isoformat())}", 553, 791, 8, "F1", (213, 226, 255), "right"),
        ]
        if page_no > 1:
            o.append(pdf_text(f"Page {page_no}", 553, 766, 8, "F1", muted, "right"))
        return o

    def add_footer(o, page_no):
        o.extend([
            pdf_line(margin, 34, margin + width, 34, border),
            pdf_text("This statement is system-generated and reflects transactions recorded at the time of issue.", margin, 20, 7.5, "F1", muted),
            pdf_text(f"Page {page_no}", margin + width, 20, 7.5, "F1", muted, "right"),
        ])

    def section_title(o, title, y):
        o.append(pdf_text(title.upper(), margin, y, 9, "F2", blue))
        o.append(pdf_line(margin, y - 7, margin + width, y - 7, border))
        return y - 20

    def info_pair(o, label, value, x, y, w=222):
        o.append(pdf_text(label.upper(), x, y, 6.8, "F2", muted, max_chars=24))
        o.append(pdf_text(value or "-", x, y - 13, 9, "F1", blue, max_chars=max(18, int(w / 4.6))))

    ops.extend(add_header())
    y = 748
    ops.append(pdf_rect(margin, y - 78, width, 78, fill=soft, stroke=border))
    info_pair(ops, "Borrower", loan.get("member_name"), margin + 14, y - 24)
    info_pair(ops, "Phone", loan.get("member_phone"), margin + 280, y - 24)
    info_pair(ops, "Address", loan.get("member_address"), margin + 14, y - 56)
    info_pair(ops, "Customer Type", "External Borrower" if loan.get("member_type") == "borrower" else "Member", margin + 280, y - 56)

    y = 640
    cards = [
        ("Principal", pdf_money(loan.get("amount"))),
        ("Total Repayable", pdf_money(summary.get("total_repayable"))),
        ("Total Paid", pdf_money(summary.get("total_paid"))),
        ("Outstanding", pdf_money(summary.get("outstanding"))),
    ]
    card_w = (width - 18) / 4
    for i, (label, value) in enumerate(cards):
        x = margin + i * (card_w + 6)
        ops.append(pdf_rect(x, y - 58, card_w, 58, fill=(255, 255, 255), stroke=border))
        ops.append(pdf_text(label.upper(), x + 10, y - 19, 6.8, "F2", muted, max_chars=18))
        ops.append(pdf_text(value, x + 10, y - 39, 10.5, "F2", blue, max_chars=18))

    y = 548
    y = section_title(ops, "Loan Details", y)
    details = [
        ("Loan ID", loan_id),
        ("Status", loan.get("status")),
        ("Annual Rate", f"{loan.get('annual_rate')}% p.a."),
        ("Term", f"{loan.get('term_months')} months"),
        ("Penalties", pdf_money(summary.get("penalties"))),
        ("Statement Date", pdf_date(date.today().isoformat())),
    ]
    for idx, (label, value) in enumerate(details):
        x = margin + (idx % 3) * 170
        row_y = y - (idx // 3) * 34
        info_pair(ops, label, value, x, row_y, 150)

    y = 444
    y = section_title(ops, "Repayment Schedule", y)
    cols = [
        ("#", margin + 8, 28, "left"),
        ("Due Date", margin + 42, 78, "left"),
        ("Principal", margin + 156, 74, "right"),
        ("Interest", margin + 245, 64, "right"),
        ("Payment", margin + 330, 74, "right"),
        ("Balance", margin + 438, 74, "right"),
    ]
    def table_header(o, y_pos):
        o.append(pdf_rect(margin, y_pos - 18, width, 18, fill=(235, 240, 247), stroke=border))
        for label, x, _, align in cols:
            o.append(pdf_text(label, x, y_pos - 12, 7.5, "F2", blue, align))
        return y_pos - 18

    page_no = 1
    y = table_header(ops, y)
    row_h = 18
    schedule = data["schedule"]
    if schedule:
        for row in schedule:
            if y < 82:
                add_footer(ops, page_no)
                pages.append("\n".join(ops))
                page_no += 1
                ops = add_header(page_no)
                y = 752
                y = section_title(ops, "Repayment Schedule Continued", y)
                y = table_header(ops, y)
            fill = (255, 255, 255) if int(row.get("installment") or 0) % 2 else (250, 252, 255)
            ops.append(pdf_rect(margin, y - row_h, width, row_h, fill=fill, stroke=border, line_width=0.35))
            values = [
                row.get("installment"),
                pdf_date(row.get("due_date")),
                f"{float(row.get('principal') or 0):,.2f}",
                f"{float(row.get('interest') or 0):,.2f}",
                f"{float(row.get('repayment') or 0):,.2f}",
                f"{float(row.get('balance') or 0):,.2f}",
            ]
            for value, (_, x, _, align) in zip(values, cols):
                ops.append(pdf_text(value, x, y - 12, 7.5, "F1", blue, align, max_chars=16))
            y -= row_h
    else:
        ops.append(pdf_rect(margin, y - 28, width, 28, fill=(255, 255, 255), stroke=border))
        ops.append(pdf_text("Schedule will be available after approval.", margin + 10, y - 18, 8, "F1", muted))
        y -= 34

    y -= 22
    if y < 170:
        add_footer(ops, page_no)
        pages.append("\n".join(ops))
        page_no += 1
        ops = add_header(page_no)
        y = 752
    y = section_title(ops, "Repayments Received", y)
    repay_cols = [
        ("Date", margin + 8, "left"),
        ("Reference", margin + 124, "left"),
        ("Method", margin + 314, "left"),
        ("Amount", margin + 500, "right"),
    ]
    ops.append(pdf_rect(margin, y - 18, width, 18, fill=(235, 240, 247), stroke=border))
    for label, x, align in repay_cols:
        ops.append(pdf_text(label, x, y - 12, 7.5, "F2", blue, align))
    y -= 18
    if data["repays"]:
        for idx, row in enumerate(data["repays"]):
            if y < 82:
                add_footer(ops, page_no)
                pages.append("\n".join(ops))
                page_no += 1
                ops = add_header(page_no)
                y = 752
                y = section_title(ops, "Repayments Received Continued", y)
            fill = (255, 255, 255) if idx % 2 == 0 else (250, 252, 255)
            ops.append(pdf_rect(margin, y - row_h, width, row_h, fill=fill, stroke=border, line_width=0.35))
            repay_values = [
                pdf_date(row.get("payment_date")),
                row.get("reference") or row.get("id"),
                row.get("method"),
                f"{float(row.get('amount') or 0):,.2f}",
            ]
            for value, (_, x, align) in zip(repay_values, repay_cols):
                ops.append(pdf_text(value, x, y - 12, 7.5, "F1", blue, align, max_chars=30))
            y -= row_h
    else:
        ops.append(pdf_rect(margin, y - 28, width, 28, fill=(255, 255, 255), stroke=border))
        ops.append(pdf_text("No repayments recorded.", margin + 10, y - 18, 8, "F1", muted))

    add_footer(ops, page_no)
    pages.append("\n".join(ops))
    return make_pdf_from_pages(pages, f"Loan Statement {loan_id}", settings.get("logo_image") or "")

def next_loan_id(db) -> str:
    row = db.execute("""
        SELECT COALESCE(MAX(CAST(SUBSTR(id, 3) AS INTEGER)), 0) AS last_no
        FROM loans
        WHERE id GLOB 'LN[0-9]*'
    """).fetchone()
    return f"LN{int(row['last_no']) + 1:03d}"

def clean_date(value, fallback=None) -> str:
    raw = (value or fallback or date.today().isoformat())
    raw = str(raw).strip()[:10]
    try:
        return date.fromisoformat(raw).isoformat()
    except ValueError:
        return date.today().isoformat()

def refresh_loan_statuses(db) -> None:
    today = date.today().isoformat()
    db.execute(
        """UPDATE loans
           SET status='overdue'
           WHERE status='active'
             AND EXISTS (
                 SELECT 1 FROM loan_schedule s
                 WHERE s.loan_id=loans.id AND s.paid=0 AND s.due_date < ?
             )""",
        (today,),
    )
    db.execute(
        """UPDATE loans
           SET status='active'
           WHERE status='overdue'
             AND NOT EXISTS (
                 SELECT 1 FROM loan_schedule s
                 WHERE s.loan_id=loans.id AND s.paid=0 AND s.due_date < ?
             )""",
        (today,),
    )

def allocate_repayment_to_schedule(db, loan_id: str, paid_date: str | None = None) -> None:
    paid_date = clean_date(paid_date)
    total_paid = db.execute(
        "SELECT COALESCE(SUM(amount),0) FROM repayments WHERE loan_id=?",
        (loan_id,),
    ).fetchone()[0]
    db.execute("UPDATE loan_schedule SET paid=0, paid_date=NULL WHERE loan_id=?", (loan_id,))
    rows = rows_to_list(db.execute(
        "SELECT id,repayment FROM loan_schedule WHERE loan_id=? ORDER BY installment",
        (loan_id,),
    ).fetchall())
    remaining = float(total_paid or 0)
    for row in rows:
        due = float(row["repayment"] or 0)
        if remaining + 0.01 >= due:
            db.execute("UPDATE loan_schedule SET paid=1, paid_date=? WHERE id=?", (paid_date, row["id"]))
            remaining -= due
        else:
            break
    total_repayable = db.execute(
        "SELECT COALESCE(SUM(repayment),0) FROM loan_schedule WHERE loan_id=?",
        (loan_id,),
    ).fetchone()[0]
    loan = row_to_dict(db.execute("SELECT * FROM loans WHERE id=?", (loan_id,)).fetchone())
    if loan:
        status = loan["status"]
        if total_repayable and float(total_paid or 0) + 0.01 >= float(total_repayable):
            status = "completed"
        elif status == "completed" and float(total_paid or 0) + 0.01 < float(total_repayable):
            status = "active"
        db.execute("UPDATE loans SET total_paid=?, status=? WHERE id=?", (total_paid, status, loan_id))
        refresh_loan_statuses(db)

def loan_risk_snapshot(db, loan_id: str) -> dict:
    today = date.today().isoformat()
    row = row_to_dict(db.execute(
        """SELECT
             COALESCE(SUM(CASE WHEN paid=0 AND due_date < ? THEN repayment ELSE 0 END),0) AS amount_in_arrears,
             COALESCE(COUNT(CASE WHEN paid=0 AND due_date < ? THEN 1 END),0) AS overdue_installments,
             MIN(CASE WHEN paid=0 AND due_date < ? THEN due_date END) AS oldest_due_date,
             MIN(CASE WHEN paid=0 THEN due_date END) AS next_due_date
           FROM loan_schedule WHERE loan_id=?""",
        (today, today, today, loan_id),
    ).fetchone())
    oldest = row.get("oldest_due_date")
    days = 0
    if oldest:
        try:
            days = max(0, (date.today() - date.fromisoformat(oldest[:10])).days)
        except Exception:
            days = 0
    return {
        "amount_in_arrears": float(row.get("amount_in_arrears") or 0),
        "overdue_installments": int(row.get("overdue_installments") or 0),
        "oldest_due_date": oldest,
        "next_due_date": row.get("next_due_date"),
        "days_in_arrears": days,
    }

def rebuild_loan_schedule(db, loan: dict, start_date: str) -> None:
    db.execute("DELETE FROM loan_schedule WHERE loan_id=?", (loan["id"],))
    schedule = build_schedule(
        loan["id"], loan["amount"], loan["annual_rate"],
        loan["term_months"], loan["method"], start_date
    )
    for row in schedule:
        db.execute(
            "INSERT INTO loan_schedule (loan_id,installment,due_date,principal,interest,repayment,balance) VALUES (?,?,?,?,?,?,?)",
            (row["loan_id"], row["installment"], row["due_date"],
             row["principal"], row["interest"], row["repayment"], row["balance"])
        )

    if not schedule:
        return
    allocate_repayment_to_schedule(db, loan["id"], start_date)

# ── Serve Frontend ────────────────────────────────────────────────────────────
@app.route("/")
def index():
    return send_from_directory(FRONTEND_DIR, "index.html")


@app.route("/v3", defaults={"path": ""})
@app.route("/v3/<path:path>")
def index_v3(path):
    if not os.path.isdir(REACT_DIST_DIR):
        return (
            "<h2>Frontend v3 build not found.</h2><p>Run <code>cd frontend &amp;&amp; npm install &amp;&amp; npm run build</code> then reload <code>/v3</code>.</p>",
            503,
            {"Content-Type": "text/html; charset=utf-8"},
        )

    target = os.path.join(REACT_DIST_DIR, path) if path else ""
    if path and os.path.isfile(target):
        return send_from_directory(REACT_DIST_DIR, path)
    return send_from_directory(REACT_DIST_DIR, "index.html")

# ══════════════════════════════════════════════════════════════════════════════
# AUTH
# ══════════════════════════════════════════════════════════════════════════════
@app.route("/api/auth/login", methods=["POST"])
def login():
    data  = request.json or {}
    username = (data.get("username") or data.get("email") or "").strip().lower()
    pwd   = data.get("password","")
    if not username or not pwd:
        return error("Username and password are required")
    attempt_key = _login_attempt_key(username)
    blocked_for = _login_block_remaining_seconds(attempt_key)
    if blocked_for > 0:
        wait_mins = max(1, -(-blocked_for // 60))
        return error(f"Too many login attempts. Try again in {wait_mins} minute(s).", 429)

    db   = get_db()
    user = row_to_dict(db.execute(
        "SELECT * FROM users WHERE (lower(username)=? OR lower(email)=?) AND active=1", (username, username)
    ).fetchone())
    db.close()

    if not user or user["password"] != hash_password(pwd):
        _record_failed_login(attempt_key)
        return error("Invalid credentials", 401)

    _clear_login_attempts(attempt_key)
    token = generate_token(user)
    return success({
        "token": token,
        "user":  {k: user[k] for k in ("id","name","username","email","role")},
    }, "Login successful")


@app.route("/api/auth/me", methods=["GET"])
@login_required
def me():
    db   = get_db()
    user = row_to_dict(db.execute(
        "SELECT id,name,email,role,created_at FROM users WHERE id=?",
        (g.user["sub"],)
    ).fetchone())
    db.close()
    return success(user)


@app.route("/api/auth/change-password", methods=["POST"])
@login_required
def change_password():
    data    = request.json or {}
    old_pwd = data.get("old_password","")
    new_pwd = data.get("new_password","")
    if not old_pwd or not new_pwd:
        return error("Both old and new password required")
    pwd_err = validate_password_strength(new_pwd)
    if pwd_err:
        return error(pwd_err)

    db   = get_db()
    user = row_to_dict(db.execute("SELECT * FROM users WHERE id=?", (g.user["sub"],)).fetchone())
    if user["password"] != hash_password(old_pwd):
        db.close()
        return error("Current password is incorrect", 401)
    db.execute("UPDATE users SET password=? WHERE id=?", (hash_password(new_pwd), g.user["sub"]))
    db.commit(); db.close()
    audit("Changed password", "Auth")
    return success(msg="Password updated")

# ══════════════════════════════════════════════════════════════════════════════
# DASHBOARD
# ══════════════════════════════════════════════════════════════════════════════
@app.route("/api/dashboard", methods=["GET"])
@login_required
def dashboard():
    db = get_db()
    refresh_loan_statuses(db)
    db.commit()

    active_loans   = db.execute("SELECT COUNT(*) FROM loans WHERE status='active'").fetchone()[0]
    overdue_loans  = db.execute("SELECT COUNT(*) FROM loans WHERE status='overdue'").fetchone()[0]
    approved_loans = db.execute("SELECT COUNT(*) FROM loans WHERE status='approved'").fetchone()[0]
    pending_loans  = db.execute("SELECT COUNT(*) FROM loans WHERE status='pending'").fetchone()[0]
    total_members  = db.execute("SELECT COUNT(*) FROM members WHERE member_type='member'").fetchone()[0]
    active_members = db.execute("SELECT COUNT(*) FROM members WHERE member_type='member' AND status='active'").fetchone()[0]
    total_savings  = db.execute("SELECT COALESCE(SUM(savings),0) FROM members WHERE member_type='member'").fetchone()[0]
    total_disbursed= db.execute("SELECT COALESCE(SUM(amount),0) FROM loans WHERE disbursed_date IS NOT NULL").fetchone()[0]
    total_repaid   = db.execute("SELECT COALESCE(SUM(amount),0) FROM repayments").fetchone()[0]
    total_penalties= db.execute("SELECT COALESCE(SUM(penalties),0) FROM loans").fetchone()[0]
    total_expenses = db.execute("SELECT COALESCE(SUM(amount),0) FROM expense_transactions").fetchone()[0]
    account_report = build_monthly_account_report(db)
    account_opening_balance = get_account_opening_balance(db)
    account_savings_collections = float(account_report["totals"]["savings_collections"] or 0)
    account_loan_repayments = float(account_report["totals"]["loan_repayments"] or 0)
    account_loan_disbursed = float(account_report["totals"]["loan_disbursed"] or 0)
    account_expenses = float(account_report["totals"]["expenses"] or 0)
    account_total_inflow = float(account_report["totals"]["inflow"] or 0)
    account_total_outflow = float(account_report["totals"]["outflow"] or 0)
    account_current_balance = account_opening_balance
    portfolio = db.execute("""
        SELECT
          COALESCE(SUM(total_repayable),0) AS total_repayable,
          COALESCE(SUM(outstanding),0) AS outstanding_portfolio,
          COALESCE(SUM(amount_in_arrears),0) AS amount_in_arrears,
          COALESCE(SUM(CASE WHEN days_in_arrears >= 30 THEN outstanding ELSE 0 END),0) AS par30_amount
        FROM (
          SELECT l.id,
                 COALESCE(s.total_repayable, l.amount) AS total_repayable,
                 MAX(COALESCE(s.total_repayable, l.amount) - l.total_paid, 0) AS outstanding,
                 COALESCE(s.amount_in_arrears,0) AS amount_in_arrears,
                 s.days_in_arrears AS days_in_arrears
          FROM loans l
          LEFT JOIN (
            SELECT loan_id,
                   SUM(repayment) AS total_repayable,
                   SUM(CASE WHEN paid=0 AND due_date < date('now') THEN repayment ELSE 0 END) AS amount_in_arrears,
                   CAST(julianday(date('now')) - julianday(MIN(CASE WHEN paid=0 AND due_date < date('now') THEN due_date END)) AS INTEGER) AS days_in_arrears
            FROM loan_schedule
            GROUP BY loan_id
          ) s ON s.loan_id=l.id
          WHERE l.disbursed_date IS NOT NULL AND l.status IN ('active','overdue','completed')
        )
    """).fetchone()
    due_today = db.execute("""
        SELECT COALESCE(SUM(repayment),0)
        FROM loan_schedule
        WHERE paid=0 AND due_date=date('now')
    """).fetchone()[0]

    # Monthly repayments (last 6 months)
    monthly = db.execute("""
        SELECT strftime('%Y-%m', payment_date) as month, SUM(amount) as total
        FROM repayments
        GROUP BY month ORDER BY month DESC LIMIT 6
    """).fetchall()

    # Recent loans with member name
    recent_loans = rows_to_list(db.execute("""
        SELECT l.id, l.amount, l.status, l.applied_date, m.name as member_name
        FROM loans l JOIN members m ON l.member_id=m.id
        ORDER BY l.created_at DESC LIMIT 5
    """).fetchall())

    # Loan status breakdown
    loan_breakdown = rows_to_list(db.execute("""
        SELECT status, COUNT(*) as count, COALESCE(SUM(amount),0) as total
        FROM loans GROUP BY status
    """).fetchall())

    # Top borrowers
    top_borrowers = rows_to_list(db.execute("""
        SELECT m.name, COUNT(l.id) as loan_count, SUM(l.amount) as total_borrowed
        FROM loans l JOIN members m ON l.member_id=m.id
        GROUP BY l.member_id ORDER BY total_borrowed DESC LIMIT 5
    """).fetchall())

    collection_rate = round((total_repaid / total_disbursed * 100), 1) if total_disbursed else 0
    par_amount = float(portfolio["amount_in_arrears"] or 0)
    outstanding_portfolio = float(portfolio["outstanding_portfolio"] or 0)
    par_rate = round((par_amount / outstanding_portfolio * 100), 1) if outstanding_portfolio else 0
    par30_amount = float(portfolio["par30_amount"] or 0)
    par30_rate = round((par30_amount / outstanding_portfolio * 100), 1) if outstanding_portfolio else 0

    db.close()
    return success({
        "stats": {
            "active_loans":    active_loans,
            "overdue_loans":   overdue_loans,
            "approved_loans":  approved_loans,
            "pending_loans":   pending_loans,
            "total_members":   total_members,
            "active_members":  active_members,
            "total_savings":   total_savings,
            "total_disbursed": total_disbursed,
            "total_repaid":    total_repaid,
            "total_penalties": total_penalties,
            "total_expenses":  total_expenses,
            "account_opening_balance": account_opening_balance,
            "account_savings_collections": account_savings_collections,
            "account_loan_repayments": account_loan_repayments,
            "account_loan_disbursed": account_loan_disbursed,
            "account_expenses": account_expenses,
            "account_total_inflow": account_total_inflow,
            "account_total_outflow": account_total_outflow,
            "account_current_balance": account_current_balance,
            "collection_rate": collection_rate,
            "outstanding_portfolio": outstanding_portfolio,
            "amount_in_arrears": par_amount,
            "portfolio_at_risk": par_rate,
            "par30_amount": par30_amount,
            "par30_rate": par30_rate,
            "due_today": due_today,
        },
        "monthly_repayments": [{"month": r["month"], "total": r["total"]} for r in reversed(monthly)],
        "recent_loans":       recent_loans,
        "loan_breakdown":     loan_breakdown,
        "top_borrowers":      top_borrowers,
    })

# ══════════════════════════════════════════════════════════════════════════════
# MEMBERS
# ══════════════════════════════════════════════════════════════════════════════
@app.route("/api/members", methods=["GET"])
@login_required
def get_members():
    q      = request.args.get("q","").strip()
    status = request.args.get("status","all")
    page   = max(1, int(request.args.get("page",1)))
    limit  = int(request.args.get("limit",10))

    where  = []
    params = []
    if q:
        where.append("(m.name LIKE ? OR m.national_id LIKE ? OR m.phone LIKE ? OR m.email LIKE ?)")
        params.extend([f"%{q}%"] * 4)
    member_type = request.args.get("type","member")
    if member_type in ("member", "borrower"):
        where.append("m.member_type=?")
        params.append(member_type)
    if status != "all":
        where.append("m.status=?")
        params.append(status)

    clause = ("WHERE " + " AND ".join(where)) if where else ""
    db     = get_db()
    total  = db.execute(f"SELECT COUNT(*) FROM members m {clause}", params).fetchone()[0]
    rows   = rows_to_list(db.execute(
        f"SELECT m.*, u.name as created_by_name FROM members m LEFT JOIN users u ON m.created_by=u.id {clause} ORDER BY m.created_at DESC LIMIT ? OFFSET ?",
        params + [limit, (page-1)*limit]
    ).fetchall())
    db.close()
    return success({"members": rows, "total": total, "page": page, "limit": limit, "pages": -(-total//limit)})


@app.route("/api/members/<member_id>", methods=["GET"])
@login_required
def get_member(member_id):
    db     = get_db()
    member = row_to_dict(db.execute("SELECT * FROM members WHERE id=?", (member_id,)).fetchone())
    if not member:
        db.close(); return error("Member not found", 404)

    loans       = rows_to_list(db.execute(
        "SELECT * FROM loans WHERE member_id=? ORDER BY created_at DESC", (member_id,)
    ).fetchall())
    savings_txn = rows_to_list(db.execute(
        "SELECT * FROM savings_transactions WHERE member_id=? ORDER BY txn_date DESC LIMIT 20", (member_id,)
    ).fetchall())
    repayments = rows_to_list(db.execute(
        "SELECT * FROM repayments WHERE member_id=? ORDER BY payment_date DESC LIMIT 30", (member_id,)
    ).fetchall())
    db.close()
    return success({"member": member, "loans": loans, "savings_transactions": savings_txn, "repayments": repayments})


@app.route("/api/members", methods=["POST"])
@login_required
@roles_required("admin","officer")
def create_member():
    d = request.json or {}
    required = ["name","joined_date"]
    if not all(d.get(k) for k in required):
        return error("Name and join date are required")

    member_type = d.get("member_type", "member")
    prefix = "B" if member_type == "borrower" else "M"
    mid = gen_id(prefix)
    national_id = d.get("national_id") or gen_id(f"{prefix}ID")
    db  = get_db()
    try:
        db.execute(
            """INSERT INTO members (id,name,phone,email,national_id,gender,dob,address,status,joined_date,savings,created_by,member_type)
               VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)""",
            (mid, d["name"], d.get("phone"), d.get("email"), national_id,
             d.get("gender"), d.get("dob"), d.get("address"), d.get("status","active"),
             d["joined_date"], 0, g.user["sub"], member_type)
        )
        db.execute("INSERT INTO savings_accounts (member_id,balance) VALUES (?,?)", (mid, 0))
        db.commit()
    except sqlite3.IntegrityError:
        db.close(); return error("National ID already exists")

    audit(f"Registered {member_type} {mid}", "Members", d["name"])
    member = row_to_dict(db.execute("SELECT * FROM members WHERE id=?", (mid,)).fetchone())
    db.close()
    return success(member, "Record registered", 201)


@app.route("/api/members/<member_id>", methods=["PUT"])
@login_required
@roles_required("admin","officer")
def update_member(member_id):
    d  = request.json or {}
    db = get_db()
    db.execute(
        """UPDATE members SET name=COALESCE(?,name), phone=COALESCE(?,phone),
           email=COALESCE(?,email), gender=COALESCE(?,gender), dob=COALESCE(?,dob),
           address=COALESCE(?,address), status=COALESCE(?,status)
           WHERE id=?""",
        (d.get("name"), d.get("phone"), d.get("email"), d.get("gender"),
         d.get("dob"), d.get("address"), d.get("status"), member_id)
    )
    db.commit()
    member = row_to_dict(db.execute("SELECT * FROM members WHERE id=?", (member_id,)).fetchone())
    db.close()
    audit(f"Updated member {member_id}", "Members")
    return success(member, "Member updated")


@app.route("/api/members/<member_id>/status", methods=["PATCH"])
@login_required
@roles_required("admin")
def toggle_member_status(member_id):
    d      = request.json or {}
    status = d.get("status","active")
    db     = get_db()
    db.execute("UPDATE members SET status=? WHERE id=?", (status, member_id))
    db.commit(); db.close()
    audit(f"Changed member {member_id} status to {status}", "Members")
    return success(msg=f"Member status set to {status}")


@app.route("/api/members/<member_id>", methods=["DELETE"])
@login_required
@roles_required("admin")
def delete_member(member_id):
    db = get_db()
    member = row_to_dict(db.execute("SELECT * FROM members WHERE id=?", (member_id,)).fetchone())
    if not member:
        db.close(); return error("Record not found", 404)
    loan_ids = [r["id"] for r in db.execute("SELECT id FROM loans WHERE member_id=?", (member_id,)).fetchall()]
    try:
        for lid in loan_ids:
            db.execute("DELETE FROM loan_schedule WHERE loan_id=?", (lid,))
            db.execute("DELETE FROM repayments WHERE loan_id=?", (lid,))
        db.execute("DELETE FROM loans WHERE member_id=?", (member_id,))
        db.execute("DELETE FROM savings_transactions WHERE member_id=?", (member_id,))
        db.execute("DELETE FROM savings_accounts WHERE member_id=?", (member_id,))
        db.execute("DELETE FROM guarantors WHERE member_id=? OR guarantor_id=?", (member_id, member_id))
        db.execute("DELETE FROM members WHERE id=?", (member_id,))
        db.commit()
    except sqlite3.IntegrityError:
        db.close(); return error("Unable to delete record with linked activity", 409)
    db.close()
    audit(f"Deleted {member.get('member_type','member')} {member_id}", "Members", member.get("name",""))
    return success(msg="Record deleted")


@app.route("/api/borrowers", methods=["GET"])
@login_required
def get_borrowers():
    q      = request.args.get("q","").strip()
    status = request.args.get("status","all")
    page   = max(1, int(request.args.get("page",1)))
    limit  = int(request.args.get("limit",10))
    where  = ["m.member_type='borrower'"]
    params = []
    if q:
        where.append("(m.name LIKE ? OR m.national_id LIKE ? OR m.phone LIKE ? OR m.email LIKE ?)")
        params.extend([f"%{q}%"] * 4)
    if status != "all":
        where.append("m.status=?")
        params.append(status)
    clause = "WHERE " + " AND ".join(where)
    db     = get_db()
    total  = db.execute(f"SELECT COUNT(*) FROM members m {clause}", params).fetchone()[0]
    rows   = rows_to_list(db.execute(
        f"SELECT m.*, u.name as created_by_name FROM members m LEFT JOIN users u ON m.created_by=u.id {clause} ORDER BY m.created_at DESC LIMIT ? OFFSET ?",
        params + [limit, (page-1)*limit]
    ).fetchall())
    db.close()
    return success({"borrowers": rows, "total": total, "page": page, "limit": limit, "pages": -(-total//limit)})


@app.route("/api/borrowers", methods=["POST"])
@login_required
@roles_required("admin","officer")
def create_borrower():
    d = request.json or {}
    d = {**d, "member_type": "borrower", "joined_date": d.get("joined_date") or date.today().isoformat()}
    if not d.get("name"):
        return error("Name is required")
    bid = gen_id("B")
    national_id = d.get("national_id") or gen_id("BID")
    db = get_db()
    try:
        db.execute(
            """INSERT INTO members (id,name,phone,email,national_id,gender,dob,address,status,joined_date,savings,created_by,member_type)
               VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)""",
            (bid, d["name"], d.get("phone"), d.get("email"), national_id,
             d.get("gender"), d.get("dob"), d.get("address"), d.get("status","active"),
             d["joined_date"], 0, g.user["sub"], "borrower")
        )
        db.execute("INSERT INTO savings_accounts (member_id,balance) VALUES (?,?)", (bid, 0))
        db.commit()
    except sqlite3.IntegrityError:
        db.close(); return error("Borrower already exists")
    borrower = row_to_dict(db.execute("SELECT * FROM members WHERE id=?", (bid,)).fetchone())
    db.close()
    audit(f"Registered borrower {bid}", "Borrowers", d["name"])
    return success(borrower, "Borrower registered", 201)

# ══════════════════════════════════════════════════════════════════════════════
# LOAN PRODUCTS
# ══════════════════════════════════════════════════════════════════════════════
@app.route("/api/loan-products", methods=["GET"])
@login_required
def get_loan_products():
    db   = get_db()
    rows = rows_to_list(db.execute("SELECT * FROM loan_products WHERE active=1").fetchall())
    db.close()
    return success(rows)

# ══════════════════════════════════════════════════════════════════════════════
# LOANS
# ══════════════════════════════════════════════════════════════════════════════
@app.route("/api/loans", methods=["GET"])
@login_required
def get_loans():
    q      = request.args.get("q","").strip()
    status = request.args.get("status","all")
    page   = max(1, int(request.args.get("page",1)))
    limit  = int(request.args.get("limit",10))

    where  = []
    params = []
    if q:
        where.append("(l.id LIKE ? OR l.member_id LIKE ? OR m.name LIKE ? OR l.purpose LIKE ?)")
        params.extend([f"%{q}%"] * 4)
    if status == "open":
        # Backward compatibility for old clients that still send status=open.
        status = "all"
    if status != "all":
        where.append("l.status=?")
        params.append(status)

    clause = ("WHERE " + " AND ".join(where)) if where else ""
    db     = get_db()
    refresh_loan_statuses(db)
    db.commit()
    total  = db.execute(
        f"SELECT COUNT(*) FROM loans l JOIN members m ON l.member_id=m.id {clause}", params
    ).fetchone()[0]
    rows   = rows_to_list(db.execute(
        f"""SELECT l.*, m.name as member_name, m.phone as member_phone,
            u.name as officer_name,
            COALESCE(l.disbursed_date, l.approved_date, l.applied_date) as borrowed_date,
            risk.next_due_date,
            COALESCE(risk.amount_in_arrears,0) as amount_in_arrears,
            COALESCE(risk.overdue_installments,0) as overdue_installments,
            risk.oldest_due_date,
            COALESCE(CAST(julianday(date('now')) - julianday(risk.oldest_due_date) AS INTEGER),0) as days_in_arrears,
            MAX(COALESCE(risk.total_repayable, l.amount) - l.total_paid, 0) as outstanding
            FROM loans l
            JOIN members m ON l.member_id=m.id
            LEFT JOIN users u ON l.officer_id=u.id
            LEFT JOIN (
                SELECT loan_id,
                       SUM(repayment) as total_repayable,
                       MIN(CASE WHEN paid=0 THEN due_date END) as next_due_date,
                       MIN(CASE WHEN paid=0 AND due_date < date('now') THEN due_date END) as oldest_due_date,
                       SUM(CASE WHEN paid=0 AND due_date < date('now') THEN repayment ELSE 0 END) as amount_in_arrears,
                       COUNT(CASE WHEN paid=0 AND due_date < date('now') THEN 1 END) as overdue_installments
                FROM loan_schedule
                GROUP BY loan_id
            ) risk ON risk.loan_id=l.id
            {clause} ORDER BY l.created_at DESC LIMIT ? OFFSET ?""",
        params + [limit, (page-1)*limit]
    ).fetchall())
    db.close()
    return success({"loans": rows, "total": total, "page": page, "limit": limit, "pages": -(-total//limit)})


@app.route("/api/loans/<loan_id>", methods=["GET"])
@login_required
def get_loan(loan_id):
    db   = get_db()
    refresh_loan_statuses(db)
    db.commit()
    loan = row_to_dict(db.execute(
        """SELECT l.*, m.name as member_name, m.phone as member_phone, m.email as member_email,
                  COALESCE(l.disbursed_date, l.approved_date, l.applied_date) as borrowed_date
           FROM loans l JOIN members m ON l.member_id=m.id
           WHERE l.id=?""", (loan_id,)
    ).fetchone())
    if not loan:
        db.close(); return error("Loan not found", 404)

    schedule = rows_to_list(db.execute(
        "SELECT * FROM loan_schedule WHERE loan_id=? ORDER BY installment", (loan_id,)
    ).fetchall())
    repays = rows_to_list(db.execute(
        "SELECT * FROM repayments WHERE loan_id=? ORDER BY payment_date DESC", (loan_id,)
    ).fetchall())
    risk = loan_risk_snapshot(db, loan_id)
    db.close()

    summary = loan_summary(loan, schedule)
    return success({"loan": loan, "schedule": schedule, "repayments": repays, "summary": summary, "risk": risk})


@app.route("/api/loans/<loan_id>/statement", methods=["GET"])
@login_required
def loan_statement(loan_id):
    data = get_loan_statement_data(loan_id)
    if not data:
        return error("Loan not found", 404)
    loan = data["loan"]
    schedule = data["schedule"]
    repays = data["repays"]
    settings = data["settings"]
    summary = data["summary"]
    schedule_rows = "".join(
        f"<tr><td>{r['installment']}</td><td>{escape(str(r['due_date']))}</td><td>{r['principal']:,.2f}</td><td>{r['interest']:,.2f}</td><td>{r['repayment']:,.2f}</td><td>{r['balance']:,.2f}</td></tr>"
        for r in schedule
    ) or "<tr><td colspan='6'>Schedule will be available after approval.</td></tr>"
    repay_rows = "".join(
        f"<tr><td>{escape(str(r['payment_date']))}</td><td>{escape(str(r.get('reference') or r['id']))}</td><td>{escape(str(r['method']))}</td><td>{r['amount']:,.2f}</td></tr>"
        for r in repays
    ) or "<tr><td colspan='4'>No repayments recorded.</td></tr>"
    logo_image = settings.get("logo_image") or ""
    logo_url = settings.get("logo_url") or ""
    logo_src = logo_image or logo_url
    logo = f"<img src='{escape(logo_src)}' alt='Logo'/>" if logo_src else f"<div class='mark'>{escape(settings.get('logo_text') or 'SF')}</div>"
    html = f"""<!doctype html><html><head><meta charset='utf-8'><title>Loan Statement {escape(loan_id)}</title>
    <style>
    @page {{ size: A5; margin: 10mm; }}
    body {{ font-family: Arial, sans-serif; color:#172033; font-size:10px; }}
    .head {{ display:flex; justify-content:space-between; align-items:flex-start; border-bottom:2px solid #172033; padding-bottom:8px; margin-bottom:10px; }}
    .brand {{ display:flex; gap:8px; align-items:center; }}
    .mark {{ width:34px; height:34px; background:#172033; color:white; display:flex; align-items:center; justify-content:center; font-weight:700; border-radius:6px; }}
    img {{ max-width:42px; max-height:42px; object-fit:contain; }}
    h1 {{ font-size:15px; margin:0; }} h2 {{ font-size:11px; margin:10px 0 5px; }}
    .muted {{ color:#64748b; }} .grid {{ display:grid; grid-template-columns:1fr 1fr; gap:6px; }}
    .box {{ border:1px solid #d7dde8; padding:6px; border-radius:5px; }}
    table {{ width:100%; border-collapse:collapse; margin-top:5px; }} th,td {{ border:1px solid #d7dde8; padding:4px; text-align:left; }}
    th {{ background:#eef2f7; }} .right {{ text-align:right; }} .print {{ position:fixed; right:12px; top:12px; }}
    @media print {{ .print {{ display:none; }} }}
    </style></head><body><button class='print' onclick='print()'>Print / Save PDF</button>
    <div class='head'><div class='brand'>{logo}<div><h1>{escape(settings.get('sacco_name') or 'SACCOFinance')}</h1>
    <div class='muted'>{escape(settings.get('address') or '')}</div><div class='muted'>{escape(settings.get('phone') or '')}</div></div></div>
    <div class='right'><strong>A5 Loan Statement</strong><div>{date.today().isoformat()}</div><div>{escape(loan_id)}</div></div></div>
    <div class='grid'>
    <div class='box'><strong>{'External Borrower' if loan.get('member_type') == 'borrower' else 'Member'}</strong><br>{escape(loan['member_name'])}<br>{escape(loan.get('member_phone') or '')}<br>{escape(loan.get('member_address') or '')}</div>
    <div class='box'><strong>Loan Summary</strong><br>Principal: KES {loan['amount']:,.2f}<br>Rate: {loan['annual_rate']}% p.a.<br>Term: {loan['term_months']} months<br>Status: {escape(loan['status'])}</div>
    <div class='box'>Total Repayable<br><strong>KES {summary['total_repayable']:,.2f}</strong></div>
    <div class='box'>Outstanding<br><strong>KES {summary['outstanding']:,.2f}</strong></div>
    </div>
    <h2>Repayment Schedule</h2><table><thead><tr><th>#</th><th>Due</th><th>Principal</th><th>Interest</th><th>Payment</th><th>Balance</th></tr></thead><tbody>{schedule_rows}</tbody></table>
    <h2>Repayments</h2><table><thead><tr><th>Date</th><th>Reference</th><th>Method</th><th>Amount</th></tr></thead><tbody>{repay_rows}</tbody></table>
    <script>setTimeout(()=>window.print(),300)</script></body></html>"""
    return make_response(html)


@app.route("/api/loans/<loan_id>/statement.pdf", methods=["GET"])
@login_required
def loan_statement_pdf(loan_id):
    data = get_loan_statement_data(loan_id)
    if not data:
        return error("Loan not found", 404)
    pdf = build_statement_pdf(loan_id, data)
    filename = f"loan-statement-{loan_id}.pdf"
    return send_file(
        BytesIO(pdf),
        mimetype="application/pdf",
        as_attachment=True,
        download_name=filename,
    )


@app.route("/api/loans", methods=["POST"])
@login_required
@roles_required("admin","officer")
def create_loan():
    d = request.json or {}
    required = ["member_id","amount","annual_rate","term_months","applied_date"]
    if not all(d.get(k) for k in required):
        return error("Required: member_id, amount, annual_rate, term_months, applied_date")
    borrowed_date = clean_date(d.get("applied_date"))

    db  = get_db()
    member = db.execute("SELECT id FROM members WHERE id=? AND status='active'", (d["member_id"],)).fetchone()
    if not member:
        db.close(); return error("Member not found or inactive")

    topup_loan = row_to_dict(db.execute(
        """SELECT * FROM loans
           WHERE member_id=? AND status IN ('pending','active','overdue')
           ORDER BY created_at ASC LIMIT 1""",
        (d["member_id"],)
    ).fetchone())

    if topup_loan:
        added_amount = float(d["amount"])
        if added_amount <= 0:
            db.close(); return error("Top-up amount must be greater than zero")

        new_amount = float(topup_loan["amount"]) + added_amount
        new_status = topup_loan["status"]
        db.execute(
            """UPDATE loans
               SET amount=?, annual_rate=?, term_months=?, method=?,
                   purpose=?, applied_date=?, disbursed_date=?, officer_id=?
               WHERE id=?""",
            (new_amount, float(d["annual_rate"]), int(d["term_months"]),
             d.get("method", topup_loan.get("method") or "reducing"),
             d.get("purpose", topup_loan.get("purpose")),
             borrowed_date, borrowed_date if topup_loan["status"] in ("active", "overdue") else topup_loan.get("disbursed_date"),
             g.user["sub"], topup_loan["id"])
        )
        loan = row_to_dict(db.execute("SELECT * FROM loans WHERE id=?", (topup_loan["id"],)).fetchone())

        if loan["status"] in ("active", "overdue"):
            start_date = borrowed_date
            rebuild_loan_schedule(db, loan, start_date)
            total_repayable = db.execute(
                "SELECT COALESCE(SUM(repayment),0) FROM loan_schedule WHERE loan_id=?",
                (loan["id"],)
            ).fetchone()[0]
            if float(loan.get("total_paid") or 0) >= float(total_repayable):
                new_status = "completed"
            db.execute("UPDATE loans SET status=? WHERE id=?", (new_status, loan["id"]))
            loan = row_to_dict(db.execute("SELECT * FROM loans WHERE id=?", (loan["id"],)).fetchone())
            # Top-up on an active/overdue loan represents new funds released.
            adjust_account_opening_balance(db, -added_amount)

        db.commit()
        db.close()
        audit(
            f"Added top-up to loan {loan['id']}",
            "Loans",
            f"Added KES {added_amount}; new principal KES {new_amount} for {d['member_id']}"
        )
        return success(loan, f"Top-up added to existing loan {loan['id']}")

    for _ in range(3):
        lid = next_loan_id(db)
        try:
            db.execute(
                """INSERT INTO loans (id,member_id,product_id,amount,annual_rate,term_months,method,
                   purpose,status,applied_date,officer_id)
                   VALUES (?,?,?,?,?,?,?,?,?,?,?)""",
                (lid, d["member_id"], d.get("product_id"), float(d["amount"]),
                 float(d["annual_rate"]), int(d["term_months"]), d.get("method","reducing"),
                 d.get("purpose"), "pending", borrowed_date, g.user["sub"])
            )
            break
        except sqlite3.IntegrityError:
            lid = None
    if not lid:
        db.close(); return error("Could not generate a new loan ID. Please try again.")
    db.commit()
    loan = row_to_dict(db.execute("SELECT * FROM loans WHERE id=?", (lid,)).fetchone())
    db.close()
    audit(f"Created loan application {lid}", "Loans", f"KES {d['amount']} for {d['member_id']}")
    return success(loan, "Loan application submitted", 201)


@app.route("/api/loans/<loan_id>", methods=["PUT"])
@login_required
@roles_required("admin","officer")
def update_loan_application(loan_id):
    d = request.json or {}
    required = ["member_id","amount","annual_rate","term_months","applied_date"]
    if not all(d.get(k) for k in required):
        return error("Required: member_id, amount, annual_rate, term_months, applied_date")
    borrowed_date = clean_date(d.get("applied_date"))

    db = get_db()
    loan = row_to_dict(db.execute("SELECT * FROM loans WHERE id=?", (loan_id,)).fetchone())
    if not loan:
        db.close(); return error("Loan not found", 404)
    status = (loan.get("status") or "").lower()
    if status not in {"pending", "approved", "rejected"}:
        db.close(); return error("Only pending, approved, or rejected loans can be edited")
    if loan.get("disbursed_date"):
        db.close(); return error("Disbursed loans cannot be edited")

    member = db.execute("SELECT id FROM members WHERE id=? AND status='active'", (d["member_id"],)).fetchone()
    if not member:
        db.close(); return error("Member or external borrower not found or inactive")

    # Editing an approved/rejected loan returns it to pending so it can be reviewed again.
    new_status = "pending" if status in {"approved", "rejected"} else "pending"
    approved_date = None if status in {"approved", "rejected"} else loan.get("approved_date")
    approved_by = None if status in {"approved", "rejected"} else loan.get("approved_by")

    db.execute(
        """UPDATE loans
           SET member_id=?, product_id=?, amount=?, annual_rate=?, term_months=?,
               method=?, purpose=?, applied_date=?, officer_id=?, status=?, approved_date=?, approved_by=?
           WHERE id=?""",
        (d["member_id"], d.get("product_id"), float(d["amount"]), float(d["annual_rate"]),
         int(d["term_months"]), d.get("method","reducing"), d.get("purpose"),
         borrowed_date, g.user["sub"], new_status, approved_date, approved_by, loan_id)
    )
    db.commit()
    loan = row_to_dict(db.execute("SELECT * FROM loans WHERE id=?", (loan_id,)).fetchone())
    db.close()
    audit(f"Updated loan application {loan_id}", "Loans", f"KES {d['amount']} for {d['member_id']}")
    if status in {"approved", "rejected"}:
        return success(loan, "Loan updated and moved to pending for review")
    return success(loan, "Loan application updated")


@app.route("/api/loans/<loan_id>", methods=["DELETE"])
@login_required
@roles_required("admin","officer")
def delete_loan(loan_id):
    db = get_db()
    loan = row_to_dict(db.execute("SELECT * FROM loans WHERE id=?", (loan_id,)).fetchone())
    if not loan:
        db.close(); return error("Loan not found", 404)

    status = (loan.get("status") or "").lower()
    if status not in {"pending", "approved", "rejected"}:
        db.close(); return error("Only pending, approved, or rejected loans can be deleted")
    if loan.get("disbursed_date"):
        db.close(); return error("Disbursed loans cannot be deleted")

    repayment_count = db.execute(
        "SELECT COUNT(*) FROM repayments WHERE loan_id=?",
        (loan_id,),
    ).fetchone()[0]
    if repayment_count > 0:
        db.close(); return error("Loans with repayments cannot be deleted")

    db.execute("DELETE FROM loan_schedule WHERE loan_id=?", (loan_id,))
    db.execute("DELETE FROM repayments WHERE loan_id=?", (loan_id,))
    db.execute("DELETE FROM notifications WHERE message LIKE ?", (f"%{loan_id}%",))
    db.execute("DELETE FROM loans WHERE id=?", (loan_id,))
    db.commit()
    db.close()
    audit(f"Deleted loan {loan_id}", "Loans", f"Status was {status}")
    return success(msg="Loan deleted")


@app.route("/api/loans/<loan_id>/approve", methods=["POST"])
@login_required
@roles_required("admin")
def approve_loan(loan_id):
    db  = get_db()
    loan = row_to_dict(db.execute("SELECT * FROM loans WHERE id=?", (loan_id,)).fetchone())
    if not loan:
        db.close(); return error("Loan not found", 404)
    if loan["status"] == "approved":
        db.close(); return success(msg="Loan already approved")
    if loan["status"] != "pending":
        db.close(); return error("Only pending loans can be approved")

    payload = request.get_json(silent=True) or {}
    approved_date = clean_date(payload.get("approved_date"), date.today().isoformat())
    if int(loan["term_months"]) <= 0:
        db.close(); return error("Loan term must be at least 1 month")
    if float(loan["amount"]) <= 0:
        db.close(); return error("Loan amount must be greater than zero")
    db.execute(
        "UPDATE loans SET status='approved', approved_date=?, approved_by=? WHERE id=?",
        (approved_date, g.user["sub"], loan_id)
    )

    # Notification
    db.execute(
        "INSERT INTO notifications (user_id,type,message) VALUES (?,?,?)",
        (loan["officer_id"] or g.user["sub"], "approved", f"Loan {loan_id} approved — KES {loan['amount']:,.0f}")
    )
    db.commit(); db.close()
    audit(f"Approved loan {loan_id}", "Loans", f"KES {loan['amount']}")
    return success(msg="Loan approved. Disburse when funds are released.")


@app.route("/api/loans/<loan_id>/disburse", methods=["POST"])
@login_required
@roles_required("admin","officer","accountant")
def disburse_loan(loan_id):
    d = request.json or {}
    disbursed_date = clean_date(d.get("disbursed_date"), date.today().isoformat())
    db  = get_db()
    loan = row_to_dict(db.execute("SELECT * FROM loans WHERE id=?", (loan_id,)).fetchone())
    if not loan:
        db.close(); return error("Loan not found", 404)
    if loan["status"] == "active":
        db.close(); return success(msg="Loan already disbursed")
    if loan["status"] != "approved":
        db.close(); return error("Only approved loans can be disbursed")
    if float(loan["amount"]) <= 0 or int(loan["term_months"]) <= 0:
        db.close(); return error("Loan amount and term must be valid before disbursement")

    db.execute(
        "UPDATE loans SET status='active', disbursed_date=?, notes=COALESCE(NULLIF(?,''),notes) WHERE id=?",
        (disbursed_date, d.get("reference") or d.get("notes") or "", loan_id)
    )
    adjust_account_opening_balance(db, -float(loan["amount"] or 0))
    loan_for_schedule = {**loan, "status": "active", "disbursed_date": disbursed_date}
    rebuild_loan_schedule(db, loan_for_schedule, disbursed_date)
    db.execute(
        "INSERT INTO notifications (user_id,type,message) VALUES (?,?,?)",
        (loan["officer_id"] or g.user["sub"], "disbursed", f"Loan {loan_id} disbursed — KES {loan['amount']:,.0f}")
    )
    db.commit(); db.close()
    audit(f"Disbursed loan {loan_id}", "Loans", f"KES {loan['amount']} on {disbursed_date}")
    return success(msg="Loan disbursed and repayment schedule generated")


@app.route("/api/loans/<loan_id>/reject", methods=["POST"])
@login_required
@roles_required("admin")
def reject_loan(loan_id):
    d    = request.json or {}
    db   = get_db()
    loan = row_to_dict(db.execute("SELECT * FROM loans WHERE id=?", (loan_id,)).fetchone())
    if not loan or loan["status"] != "pending":
        db.close(); return error("Loan not found or not pending")
    db.execute("UPDATE loans SET status='rejected', notes=? WHERE id=?", (d.get("reason",""), loan_id))
    db.commit(); db.close()
    audit(f"Rejected loan {loan_id}", "Loans", d.get("reason",""))
    return success(msg="Loan rejected")


@app.route("/api/loans/<loan_id>/schedule", methods=["GET"])
@login_required
def get_loan_schedule(loan_id):
    db       = get_db()
    schedule = rows_to_list(db.execute(
        "SELECT * FROM loan_schedule WHERE loan_id=? ORDER BY installment", (loan_id,)
    ).fetchall())
    db.close()
    return success(schedule)

# ══════════════════════════════════════════════════════════════════════════════
# REPAYMENTS
# ══════════════════════════════════════════════════════════════════════════════
@app.route("/api/repayments", methods=["GET"])
@login_required
def get_repayments():
    q     = request.args.get("q","").strip()
    page  = max(1, int(request.args.get("page",1)))
    limit = int(request.args.get("limit",10))

    where  = []
    params = []
    if q:
        where.append("(r.id LIKE ? OR r.loan_id LIKE ? OR m.name LIKE ? OR r.reference LIKE ?)")
        params.extend([f"%{q}%"] * 4)

    clause = ("WHERE " + " AND ".join(where)) if where else ""
    db     = get_db()
    total  = db.execute(
        f"SELECT COUNT(*) FROM repayments r JOIN members m ON r.member_id=m.id {clause}", params
    ).fetchone()[0]
    rows   = rows_to_list(db.execute(
        f"""SELECT r.*, m.name as member_name, u.name as recorded_by_name
            FROM repayments r
            JOIN members m ON r.member_id=m.id
            LEFT JOIN users u ON r.recorded_by=u.id
            {clause} ORDER BY r.payment_date DESC, r.created_at DESC LIMIT ? OFFSET ?""",
        params + [limit, (page-1)*limit]
    ).fetchall())
    db.close()
    return success({"repayments": rows, "total": total, "page": page, "limit": limit, "pages": -(-total//limit)})


@app.route("/api/repayments", methods=["POST"])
@login_required
@roles_required("admin","officer","cashier")
def create_repayment():
    d = request.json or {}
    required = ["loan_id","amount","payment_date"]
    if not all(d.get(k) for k in required):
        return error("Required: loan_id, amount, payment_date")

    amount = float(d["amount"])
    if amount <= 0:
        return error("Amount must be positive")

    db   = get_db()
    refresh_loan_statuses(db)
    db.commit()
    loan = row_to_dict(db.execute("SELECT * FROM loans WHERE id=?", (d["loan_id"],)).fetchone())
    if not loan:
        db.close(); return error("Loan not found")
    if loan["status"] not in ("active","overdue"):
        db.close(); return error("Loan is not active")

    rid = gen_id("R")
    method = d.get("method","cash")
    ref = d.get("reference") or (
        "QK" + gen_id() if method=="mpesa" else
        "BNK" + gen_id() if method=="bank" else
        "CSH" + gen_id()
    )
    if ref and db.execute("SELECT id FROM repayments WHERE reference=?", (ref,)).fetchone():
        db.close(); return error("A repayment with this reference already exists")

    schedule = rows_to_list(db.execute(
        "SELECT * FROM loan_schedule WHERE loan_id=? ORDER BY installment", (loan["id"],)
    ).fetchall())
    if not schedule:
        db.close(); return error("Loan has no repayment schedule. Disburse it first.")
    total_repayable = sum(float(r["repayment"] or 0) for r in schedule)
    outstanding = max(0, total_repayable - float(loan["total_paid"] or 0))
    if amount > outstanding + 0.01:
        db.close(); return error(f"Payment exceeds outstanding balance of KES {outstanding:,.2f}")

    db.execute(
        "INSERT INTO repayments (id,loan_id,member_id,amount,payment_date,method,reference,type,recorded_by) VALUES (?,?,?,?,?,?,?,?,?)",
        (rid, loan["id"], loan["member_id"], amount, d["payment_date"],
         method, ref, d.get("type","installment"), g.user["sub"])
    )
    adjust_account_opening_balance(db, amount)

    allocate_repayment_to_schedule(db, loan["id"], d["payment_date"])

    db.commit()
    repayment = row_to_dict(db.execute("SELECT * FROM repayments WHERE id=?", (rid,)).fetchone())
    db.close()
    audit(f"Recorded repayment {rid}", "Repayments", f"KES {amount} for loan {loan['id']}")
    return success({**repayment, "reference": ref}, "Repayment recorded", 201)

# ══════════════════════════════════════════════════════════════════════════════
# SAVINGS
# ══════════════════════════════════════════════════════════════════════════════
@app.route("/api/savings", methods=["GET"])
@login_required
def get_savings():
    db   = get_db()
    rows = rows_to_list(db.execute(
        """SELECT m.id, m.name, m.phone, m.status, sa.balance,
           COUNT(st.id) as txn_count
           FROM members m
           LEFT JOIN savings_accounts sa ON m.id=sa.member_id
           LEFT JOIN savings_transactions st ON m.id=st.member_id
           WHERE m.member_type='member'
           GROUP BY m.id ORDER BY sa.balance DESC"""
    ).fetchall())
    db.close()
    return success(rows)


@app.route("/api/savings/transactions", methods=["GET"])
@login_required
def get_savings_transactions():
    member_id = request.args.get("member_id","")
    q         = request.args.get("q","")
    page      = max(1, int(request.args.get("page",1)))
    limit     = int(request.args.get("limit",15))

    where  = []
    params = []
    if member_id:
        where.append("st.member_id=?"); params.append(member_id)
    if q:
        where.append("(m.name LIKE ? OR st.reference LIKE ?)"); params.extend([f"%{q}%"]*2)

    clause = ("WHERE " + " AND ".join(where)) if where else ""
    db     = get_db()
    total  = db.execute(
        f"SELECT COUNT(*) FROM savings_transactions st JOIN members m ON st.member_id=m.id {clause}", params
    ).fetchone()[0]
    rows   = rows_to_list(db.execute(
        f"""SELECT st.*, m.name as member_name
            FROM savings_transactions st JOIN members m ON st.member_id=m.id
            {clause} ORDER BY st.txn_date DESC, st.created_at DESC LIMIT ? OFFSET ?""",
        params + [limit, (page-1)*limit]
    ).fetchall())
    db.close()
    return success({"transactions": rows, "total": total, "page": page, "limit": limit, "pages": -(-total//limit)})


@app.route("/api/savings/deposit", methods=["POST"])
@login_required
@roles_required("admin","officer","cashier","accountant")
def savings_deposit():
    d = request.json or {}
    if not d.get("member_id") or not d.get("amount"):
        return error("member_id and amount required")

    amount = float(d["amount"])
    if amount <= 0:
        return error("Amount must be positive")

    db     = get_db()
    member = row_to_dict(db.execute("SELECT * FROM members WHERE id=?", (d["member_id"],)).fetchone())
    if not member:
        db.close(); return error("Member not found")

    new_balance = member["savings"] + amount
    tid = gen_id("ST")
    ref = d.get("reference") or "DEP" + gen_id()

    db.execute("UPDATE members SET savings=? WHERE id=?", (new_balance, d["member_id"]))
    db.execute("UPDATE savings_accounts SET balance=? WHERE member_id=?", (new_balance, d["member_id"]))
    db.execute(
        "INSERT INTO savings_transactions (id,member_id,type,amount,category,txn_date,reference,balance_after,recorded_by) VALUES (?,?,?,?,?,?,?,?,?)",
        (tid, d["member_id"], "deposit", amount, d.get("category","voluntary"),
         d.get("date", date.today().isoformat()), ref, new_balance, g.user["sub"])
    )
    adjust_account_opening_balance(db, amount)
    db.commit()
    db.close()
    audit(f"Recorded deposit {tid}", "Savings", f"KES {amount} for {d['member_id']}")
    return success({"reference": ref, "new_balance": new_balance}, "Deposit recorded", 201)


@app.route("/api/savings/withdraw", methods=["POST"])
@login_required
@roles_required("admin","officer","cashier","accountant")
def savings_withdraw():
    d = request.json or {}
    if not d.get("member_id") or not d.get("amount"):
        return error("member_id and amount required")

    amount = float(d["amount"])
    if amount <= 0:
        return error("Amount must be positive")
    db     = get_db()
    member = row_to_dict(db.execute("SELECT * FROM members WHERE id=?", (d["member_id"],)).fetchone())
    if not member:
        db.close(); return error("Member not found")
    if member["savings"] < amount:
        db.close(); return error(f"Insufficient balance. Current: KES {member['savings']:,.2f}")

    new_balance = member["savings"] - amount
    tid = gen_id("ST")
    ref = d.get("reference") or "WDR" + gen_id()

    db.execute("UPDATE members SET savings=? WHERE id=?", (new_balance, d["member_id"]))
    db.execute("UPDATE savings_accounts SET balance=? WHERE member_id=?", (new_balance, d["member_id"]))
    db.execute(
        "INSERT INTO savings_transactions (id,member_id,type,amount,category,txn_date,reference,balance_after,recorded_by) VALUES (?,?,?,?,?,?,?,?,?)",
        (tid, d["member_id"], "withdrawal", amount, d.get("category","voluntary"),
         d.get("date", date.today().isoformat()), ref, new_balance, g.user["sub"])
    )
    adjust_account_opening_balance(db, -amount)
    db.commit(); db.close()
    audit(f"Recorded withdrawal {tid}", "Savings", f"KES {amount} for {d['member_id']}")
    return success({"reference": ref, "new_balance": new_balance}, "Withdrawal processed", 201)

# ══════════════════════════════════════════════════════════════════════════════
# EXPENSES
# ══════════════════════════════════════════════════════════════════════════════
@app.route("/api/expenses/accounts", methods=["GET"])
@login_required
def get_expense_accounts():
    include_inactive = request.args.get("include_inactive", "false").strip().lower() == "true"
    db = get_db()
    if include_inactive:
        rows = rows_to_list(db.execute(
            """SELECT ea.*, u.name as created_by_name
               FROM expense_accounts ea
               LEFT JOIN users u ON ea.created_by=u.id
               ORDER BY ea.active DESC, ea.name ASC"""
        ).fetchall())
    else:
        rows = rows_to_list(db.execute(
            """SELECT ea.*, u.name as created_by_name
               FROM expense_accounts ea
               LEFT JOIN users u ON ea.created_by=u.id
               WHERE ea.active=1
               ORDER BY ea.name ASC"""
        ).fetchall())
    db.close()
    return success(rows)


@app.route("/api/expenses/accounts", methods=["POST"])
@login_required
@roles_required("admin", "accountant")
def create_expense_account():
    d = request.json or {}
    name = (d.get("name") or "").strip()
    code = (d.get("code") or "").strip().upper()
    description = (d.get("description") or "").strip()

    if not name:
        return error("Account name is required")

    db = get_db()
    try:
        db.execute(
            "INSERT INTO expense_accounts (code,name,description,active,created_by) VALUES (?,?,?,?,?)",
            (code or None, name, description or None, 1, g.user["sub"])
        )
        db.commit()
    except sqlite3.IntegrityError:
        db.close()
        return error("Expense account name or code already exists")

    account = row_to_dict(db.execute(
        "SELECT * FROM expense_accounts WHERE id=last_insert_rowid()"
    ).fetchone())
    db.close()
    audit(f"Created expense account {account['name']}", "Expenses", account.get("code") or "")
    return success(account, "Expense account created", 201)


@app.route("/api/expenses/accounts/<int:account_id>/status", methods=["PATCH"])
@login_required
@roles_required("admin", "accountant")
def update_expense_account_status(account_id):
    d = request.json or {}
    active = 1 if bool(d.get("active", True)) else 0

    db = get_db()
    account = row_to_dict(db.execute(
        "SELECT * FROM expense_accounts WHERE id=?",
        (account_id,)
    ).fetchone())
    if not account:
        db.close()
        return error("Expense account not found", 404)

    db.execute("UPDATE expense_accounts SET active=? WHERE id=?", (active, account_id))
    db.commit()
    updated = row_to_dict(db.execute("SELECT * FROM expense_accounts WHERE id=?", (account_id,)).fetchone())
    db.close()
    audit(
        f"{'Activated' if active else 'Deactivated'} expense account {updated.get('name')}",
        "Expenses",
        updated.get("code") or ""
    )
    return success(updated, "Expense account status updated")


@app.route("/api/expenses/transactions", methods=["GET"])
@login_required
def get_expense_transactions():
    q = request.args.get("q", "").strip()
    account_id = request.args.get("account_id", "").strip()
    date_from = request.args.get("date_from", "").strip()
    date_to = request.args.get("date_to", "").strip()
    page = max(1, int(request.args.get("page", 1)))
    limit = int(request.args.get("limit", 15))

    where = []
    params = []
    if q:
        where.append("(et.id LIKE ? OR et.reference LIKE ? OR et.payee LIKE ? OR et.notes LIKE ? OR ea.name LIKE ?)")
        params.extend([f"%{q}%"] * 5)
    if account_id:
        where.append("et.account_id=?")
        params.append(account_id)
    if date_from:
        where.append("et.expense_date>=?")
        params.append(clean_date(date_from))
    if date_to:
        where.append("et.expense_date<=?")
        params.append(clean_date(date_to))

    clause = ("WHERE " + " AND ".join(where)) if where else ""
    db = get_db()
    total = db.execute(
        f"""SELECT COUNT(*)
            FROM expense_transactions et
            JOIN expense_accounts ea ON et.account_id=ea.id
            {clause}""",
        params,
    ).fetchone()[0]
    total_amount = db.execute(
        f"""SELECT COALESCE(SUM(et.amount),0)
            FROM expense_transactions et
            JOIN expense_accounts ea ON et.account_id=ea.id
            {clause}""",
        params,
    ).fetchone()[0]
    rows = rows_to_list(db.execute(
        f"""SELECT et.*, ea.name as account_name, ea.code as account_code, u.name as recorded_by_name
            FROM expense_transactions et
            JOIN expense_accounts ea ON et.account_id=ea.id
            LEFT JOIN users u ON et.recorded_by=u.id
            {clause}
            ORDER BY et.expense_date DESC, et.created_at DESC
            LIMIT ? OFFSET ?""",
        params + [limit, (page - 1) * limit],
    ).fetchall())
    db.close()
    return success({
        "transactions": rows,
        "total": total,
        "total_amount": total_amount,
        "page": page,
        "limit": limit,
        "pages": -(-total // limit),
    })


@app.route("/api/expenses/transactions", methods=["POST"])
@login_required
@roles_required("admin", "accountant")
def create_expense_transaction():
    d = request.json or {}
    if not d.get("account_id") or not d.get("amount"):
        return error("account_id and amount are required")

    amount = float(d.get("amount") or 0)
    if amount <= 0:
        return error("Amount must be greater than zero")

    expense_date = clean_date(d.get("expense_date"), date.today().isoformat())
    db = get_db()
    account = row_to_dict(db.execute(
        "SELECT * FROM expense_accounts WHERE id=?",
        (d["account_id"],)
    ).fetchone())
    if not account:
        db.close()
        return error("Expense account not found")
    if not int(account.get("active") or 0):
        db.close()
        return error("Expense account is inactive")

    eid = gen_id("EX")
    reference = (d.get("reference") or "").strip() or f"EXP-{gen_id()}"
    payee = (d.get("payee") or "").strip()
    notes = (d.get("notes") or "").strip()

    db.execute(
        """INSERT INTO expense_transactions
           (id,account_id,amount,expense_date,reference,payee,notes,recorded_by)
           VALUES (?,?,?,?,?,?,?,?)""",
        (eid, int(d["account_id"]), amount, expense_date, reference, payee or None, notes or None, g.user["sub"])
    )
    adjust_account_opening_balance(db, -amount)
    db.commit()
    row = row_to_dict(db.execute(
        """SELECT et.*, ea.name as account_name, ea.code as account_code
           FROM expense_transactions et
           JOIN expense_accounts ea ON et.account_id=ea.id
           WHERE et.id=?""",
        (eid,),
    ).fetchone())
    db.close()
    audit(f"Recorded expense {eid}", "Expenses", f"KES {amount:,.2f} - {row.get('account_name')}")
    return success(row, "Expense recorded", 201)


@app.route("/api/expenses/transactions/<expense_id>", methods=["DELETE"])
@login_required
@roles_required("admin", "accountant")
def delete_expense_transaction(expense_id):
    db = get_db()
    row = row_to_dict(db.execute(
        """SELECT et.*, ea.name as account_name
           FROM expense_transactions et
           JOIN expense_accounts ea ON et.account_id=ea.id
           WHERE et.id=?""",
        (expense_id,),
    ).fetchone())
    if not row:
        db.close()
        return error("Expense not found", 404)

    db.execute("DELETE FROM expense_transactions WHERE id=?", (expense_id,))
    db.commit()
    db.close()
    audit(
        f"Deleted expense {expense_id}",
        "Expenses",
        f"KES {float(row.get('amount') or 0):,.2f} - {row.get('account_name')}"
    )
    return success(msg="Expense deleted")

# ══════════════════════════════════════════════════════════════════════════════
# REPORTS
# ══════════════════════════════════════════════════════════════════════════════
@app.route("/api/reports/portfolio", methods=["GET"])
@login_required
def report_portfolio():
    db = get_db()
    refresh_loan_statuses(db)
    db.commit()
    rows = rows_to_list(db.execute(
        """SELECT l.*, m.name as member_name,
           COALESCE(risk.total_repayable, l.amount) as total_repayable,
           MAX(COALESCE(risk.total_repayable, l.amount) - l.total_paid, 0) as outstanding,
           COALESCE(risk.amount_in_arrears,0) as amount_in_arrears,
           COALESCE(risk.overdue_installments,0) as overdue_installments,
           COALESCE(CAST(julianday(date('now')) - julianday(risk.oldest_due_date) AS INTEGER),0) as days_in_arrears,
           risk.next_due_date
           FROM loans l
           JOIN members m ON l.member_id=m.id
           LEFT JOIN (
             SELECT loan_id,
                    SUM(repayment) as total_repayable,
                    MIN(CASE WHEN paid=0 THEN due_date END) as next_due_date,
                    MIN(CASE WHEN paid=0 AND due_date < date('now') THEN due_date END) as oldest_due_date,
                    SUM(CASE WHEN paid=0 AND due_date < date('now') THEN repayment ELSE 0 END) as amount_in_arrears,
                    COUNT(CASE WHEN paid=0 AND due_date < date('now') THEN 1 END) as overdue_installments
             FROM loan_schedule
             GROUP BY loan_id
           ) risk ON risk.loan_id=l.id
           ORDER BY l.disbursed_date DESC"""
    ).fetchall())
    totals = db.execute(
        """SELECT
           COALESCE(SUM(amount),0) as total_disbursed,
           COALESCE(SUM(total_paid),0) as total_repaid,
           COALESCE(SUM(total_repayable),0) as total_repayable,
           COALESCE(SUM(MAX(total_repayable - total_paid, 0)),0) as total_outstanding,
           COALESCE(SUM(amount_in_arrears),0) as amount_in_arrears,
           COALESCE(SUM(CASE WHEN days_in_arrears >= 30 THEN MAX(total_repayable - total_paid, 0) ELSE 0 END),0) as par30_amount,
           COALESCE(SUM(penalties),0) as total_penalties,
           COUNT(*) as total_loans
           FROM (
               SELECT l.id, l.amount, l.total_paid, l.penalties, l.disbursed_date,
                       COALESCE(risk.total_repayable, l.amount) as total_repayable,
                       COALESCE(risk.amount_in_arrears,0) as amount_in_arrears,
                       risk.days_in_arrears
               FROM loans l
               LEFT JOIN (
                 SELECT loan_id,
                        SUM(repayment) as total_repayable,
                        SUM(CASE WHEN paid=0 AND due_date < date('now') THEN repayment ELSE 0 END) as amount_in_arrears,
                        CAST(julianday(date('now')) - julianday(MIN(CASE WHEN paid=0 AND due_date < date('now') THEN due_date END)) AS INTEGER) as days_in_arrears
                 FROM loan_schedule
                 GROUP BY loan_id
               ) risk ON risk.loan_id=l.id
               WHERE l.disbursed_date IS NOT NULL
           )"""
    ).fetchone()
    db.close()
    return success({"loans": rows, "totals": dict(totals)})


@app.route("/api/reports/savings", methods=["GET"])
@login_required
def report_savings():
    db = get_db()
    rows = rows_to_list(db.execute(
        """SELECT m.id, m.name, m.status, sa.balance,
           COALESCE(SUM(CASE WHEN st.type='deposit' THEN st.amount ELSE 0 END),0) as total_deposits,
           COALESCE(SUM(CASE WHEN st.type='withdrawal' THEN st.amount ELSE 0 END),0) as total_withdrawals
           FROM members m
           LEFT JOIN savings_accounts sa ON m.id=sa.member_id
           LEFT JOIN savings_transactions st ON m.id=st.member_id
           WHERE m.member_type='member'
           GROUP BY m.id ORDER BY sa.balance DESC"""
    ).fetchall())
    db.close()
    return success(rows)


@app.route("/api/reports/account-monthly", methods=["GET"])
@login_required
def report_account_monthly():
    db = get_db()
    data = build_monthly_account_report(db)
    db.close()
    return success(data)


@app.route("/api/reports/export/<report_type>", methods=["GET"])
@login_required
def export_report(report_type):
    import csv, io
    db = get_db()
    export_format = (request.args.get("format") or "csv").strip().lower()
    rows = []
    headers = []

    if report_type == "loans":
        headers = ["ID","Member","Amount","Rate%","Term","Method","Status","Disbursed","Paid","Outstanding","Penalties"]
        rows = rows_to_list(db.execute(
            """SELECT l.id,m.name,l.amount,l.annual_rate,l.term_months,l.method,l.status,l.disbursed_date,l.total_paid,
                      (COALESCE(SUM(s.repayment), l.amount)-l.total_paid) as outstanding,l.penalties
               FROM loans l
               JOIN members m ON l.member_id=m.id
               LEFT JOIN loan_schedule s ON s.loan_id=l.id
               GROUP BY l.id"""
        ).fetchall())
    elif report_type == "repayments":
        headers = ["ID","Loan","Member","Amount","Date","Method","Reference","Type"]
        rows = rows_to_list(db.execute(
            "SELECT r.id,r.loan_id,m.name,r.amount,r.payment_date,r.method,r.reference,r.type FROM repayments r JOIN members m ON r.member_id=m.id ORDER BY r.payment_date DESC"
        ).fetchall())
    elif report_type == "savings":
        headers = ["Member ID","Member","Balance","Status"]
        rows = rows_to_list(db.execute(
            "SELECT m.id,m.name,sa.balance,m.status FROM members m LEFT JOIN savings_accounts sa ON m.id=sa.member_id WHERE m.member_type='member' ORDER BY sa.balance DESC"
        ).fetchall())
    elif report_type == "expenses":
        headers = ["ID","Date","Account","Account Code","Amount","Payee","Reference","Notes","Recorded By"]
        rows = rows_to_list(db.execute(
            """SELECT et.id, et.expense_date, ea.name, ea.code, et.amount, et.payee, et.reference, et.notes, u.name
               FROM expense_transactions et
               JOIN expense_accounts ea ON et.account_id=ea.id
               LEFT JOIN users u ON et.recorded_by=u.id
               ORDER BY et.expense_date DESC, et.created_at DESC"""
        ).fetchall())
    elif report_type == "account-monthly":
        headers = ["Month","Opening Balance","Savings Collections","Loan Repayments","Total Inflow","Loans Disbursed","Expenses","Total Outflow","Net Movement","Closing Balance"]
        data = build_monthly_account_report(db)
        for row in data.get("months", []):
            rows.append({
                "Month": row.get("month"),
                "Opening Balance": row.get("opening_balance"),
                "Savings Collections": row.get("savings_collections"),
                "Loan Repayments": row.get("loan_repayments"),
                "Total Inflow": row.get("inflow"),
                "Loans Disbursed": row.get("loan_disbursed"),
                "Expenses": row.get("expenses"),
                "Total Outflow": row.get("outflow"),
                "Net Movement": row.get("net"),
                "Closing Balance": row.get("closing_balance"),
            })
    elif report_type == "members":
        headers = ["ID","Name","Phone","Email","National ID","Status","Joined","Savings"]
        rows = rows_to_list(db.execute("SELECT id,name,phone,email,national_id,status,joined_date,savings FROM members WHERE member_type='member'").fetchall())
    else:
        db.close(); return error("Unknown report type")

    # CSV output is the default and remains backward-compatible.
    if export_format not in ("csv", "xlsx"):
        db.close()
        return error("Unsupported export format. Use csv or xlsx.")

    db.close()
    from flask import Response
    audit(f"Exported {report_type} report", "Reports")
    if export_format == "xlsx":
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill
        except Exception:
            return error("Excel export dependency missing. Install openpyxl.")

        wb = Workbook()
        ws = wb.active
        ws.title = (report_type or "report")[:31]
        ws.append(headers)
        header_fill = PatternFill(start_color="E5EEF9", end_color="E5EEF9", fill_type="solid")
        for idx in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=idx)
            cell.font = Font(bold=True)
            cell.fill = header_fill

        for row in rows:
            if isinstance(row, dict):
                ws.append([row.get(h) for h in headers])
            else:
                ws.append(list(row))

        # Set readable default widths.
        for col in ws.columns:
            max_len = max((len(str(c.value or "")) for c in col), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(42, max(12, max_len + 2))

        stream = BytesIO()
        wb.save(stream)
        stream.seek(0)
        return send_file(
            stream,
            as_attachment=True,
            download_name=f"{report_type}-report.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(headers)
    for row in rows:
        writer.writerow([row.get(h) for h in headers] if isinstance(row, dict) else list(row))
    return Response(
        output.getvalue(),
        mimetype="text/csv",
        headers={"Content-Disposition": f"attachment;filename={report_type}-report.csv"}
    )


# ══════════════════════════════════════════════════════════════════════════════
# ACCOUNTING
# ══════════════════════════════════════════════════════════════════════════════
@app.route("/api/accounting/chart-of-accounts", methods=["GET"])
@login_required
def accounting_chart_of_accounts():
    return success(ACCOUNTING_ACCOUNTS)


@app.route("/api/accounting/journal-entries", methods=["GET"])
@login_required
def accounting_journal_entries():
    date_from = request.args.get("date_from", "").strip()
    date_to = request.args.get("date_to", "").strip()
    source = request.args.get("source", "").strip()
    page = max(1, int(request.args.get("page", 1)))
    limit = int(request.args.get("limit", 15))
    db = get_db()
    entries = _filter_journal(build_accounting_journal(db), date_from, date_to, source)
    db.close()
    total = len(entries)
    start = (page - 1) * limit
    return success({
        "entries": entries[start:start + limit],
        "total": total,
        "page": page,
        "limit": limit,
        "pages": -(-total // limit) if limit else 1,
    })


@app.route("/api/accounting/general-ledger", methods=["GET"])
@login_required
def accounting_general_ledger():
    account_code = request.args.get("account_code", "").strip()
    date_from = request.args.get("date_from", "").strip()
    date_to = request.args.get("date_to", "").strip()
    db = get_db()
    entries = _filter_journal(build_accounting_journal(db), date_from, date_to)
    db.close()
    rows = []
    running = 0.0
    for entry in entries:
        for line in entry["lines"]:
            if account_code and line["account_code"] != account_code:
                continue
            running += float(line.get("debit") or 0) - float(line.get("credit") or 0)
            rows.append({
                "entry_id": entry["id"],
                "entry_date": entry["entry_date"],
                "source": entry["source"],
                "source_id": entry["source_id"],
                "description": entry["description"],
                **line,
                "running_balance": round(running, 2),
            })
    return success({
        "account_code": account_code,
        "account": ACCOUNT_BY_CODE.get(account_code),
        "lines": rows,
    })


@app.route("/api/accounting/trial-balance", methods=["GET"])
@login_required
def accounting_trial_balance():
    db = get_db()
    data = build_trial_balance(
        db,
        request.args.get("date_from", "").strip(),
        request.args.get("date_to", "").strip(),
    )
    db.close()
    return success(data)


@app.route("/api/accounting/profit-loss", methods=["GET"])
@login_required
def accounting_profit_loss():
    db = get_db()
    data = build_profit_loss(
        db,
        request.args.get("date_from", "").strip(),
        request.args.get("date_to", "").strip(),
    )
    db.close()
    return success(data)


@app.route("/api/accounting/balance-sheet", methods=["GET"])
@login_required
def accounting_balance_sheet():
    db = get_db()
    data = build_balance_sheet(db, request.args.get("date_to", "").strip())
    db.close()
    return success(data)


@app.route("/api/accounting/cash-flow", methods=["GET"])
@login_required
def accounting_cash_flow():
    db = get_db()
    data = build_cash_flow_statement(
        db,
        request.args.get("date_from", "").strip(),
        request.args.get("date_to", "").strip(),
    )
    db.close()
    return success(data)

# ══════════════════════════════════════════════════════════════════════════════
# NOTIFICATIONS
# ══════════════════════════════════════════════════════════════════════════════
@app.route("/api/notifications", methods=["GET"])
@login_required
def get_notifications():
    db = get_db()
    rows = rows_to_list(db.execute(
        "SELECT * FROM notifications WHERE user_id=? OR user_id IS NULL ORDER BY created_at DESC LIMIT 30",
        (g.user["sub"],)
    ).fetchall())
    unread = db.execute(
        "SELECT COUNT(*) FROM notifications WHERE (user_id=? OR user_id IS NULL) AND read=0",
        (g.user["sub"],)
    ).fetchone()[0]
    db.close()
    return success({"notifications": rows, "unread": unread})


@app.route("/api/notifications/read-all", methods=["POST"])
@login_required
def mark_all_read():
    db = get_db()
    db.execute("UPDATE notifications SET read=1 WHERE user_id=? OR user_id IS NULL", (g.user["sub"],))
    db.commit(); db.close()
    return success(msg="All notifications marked as read")


@app.route("/api/notifications/<int:nid>/read", methods=["PATCH"])
@login_required
def mark_read(nid):
    db = get_db()
    db.execute("UPDATE notifications SET read=1 WHERE id=?", (nid,))
    db.commit(); db.close()
    return success(msg="Notification marked as read")

# ══════════════════════════════════════════════════════════════════════════════
# AUDIT LOG
# ══════════════════════════════════════════════════════════════════════════════
@app.route("/api/audit-logs", methods=["GET"])
@login_required
@roles_required("admin","accountant")
def get_audit_logs():
    module = request.args.get("module","")
    page   = max(1, int(request.args.get("page",1)))
    limit  = int(request.args.get("limit",15))
    where  = "WHERE module=?" if module else ""
    params = [module] if module else []
    db     = get_db()
    total  = db.execute(f"SELECT COUNT(*) FROM audit_logs {where}", params).fetchone()[0]
    rows   = rows_to_list(db.execute(
        f"SELECT * FROM audit_logs {where} ORDER BY created_at DESC LIMIT ? OFFSET ?",
        params + [limit, (page-1)*limit]
    ).fetchall())
    db.close()
    return success({"logs": rows, "total": total, "page": page, "limit": limit, "pages": -(-total//limit)})

# ══════════════════════════════════════════════════════════════════════════════
# USERS (admin only)
# ══════════════════════════════════════════════════════════════════════════════
@app.route("/api/users", methods=["GET"])
@roles_required("admin")
def get_users():
    db   = get_db()
    rows = rows_to_list(db.execute(
        "SELECT id,name,username,email,role,active,created_at FROM users ORDER BY id"
    ).fetchall())
    db.close()
    return success(rows)


@app.route("/api/users", methods=["POST"])
@roles_required("admin")
def create_user():
    d = request.json or {}
    if not all(d.get(k) for k in ("name","username","password","role")):
        return error("name, username, password, role required")
    username = d["username"].strip().lower()
    role = (d.get("role") or "").strip().lower()
    role_err = validate_user_role(role)
    if role_err:
        return error(role_err)
    pwd_err = validate_password_strength(d["password"])
    if pwd_err:
        return error(pwd_err)
    email = (d.get("email") or f"{username}@local.sacco").strip().lower()
    db = get_db()
    try:
        db.execute(
            "INSERT INTO users (name,username,email,password,role) VALUES (?,?,?,?,?)",
            (d["name"].strip(), username, email, hash_password(d["password"]), role)
        )
        db.commit()
    except sqlite3.IntegrityError:
        db.close(); return error("Username or email already registered")
    db.close()
    audit(f"Created user {username}", "Users")
    return success(msg="User created", code=201)


@app.route("/api/users/<int:user_id>", methods=["PUT"])
@roles_required("admin")
def update_user(user_id):
    d = request.json or {}
    if not all(d.get(k) for k in ("name","username","role")):
        return error("name, username and role required")
    username = d["username"].strip().lower()
    role = (d.get("role") or "").strip().lower()
    role_err = validate_user_role(role)
    if role_err:
        return error(role_err)
    if str(user_id) == str(g.user["sub"]) and role != "admin":
        return error("You cannot remove your own admin role")
    email = (d.get("email") or f"{username}@local.sacco").strip().lower()
    db = get_db()
    user = row_to_dict(db.execute("SELECT * FROM users WHERE id=?", (user_id,)).fetchone())
    if not user:
        db.close(); return error("User not found", 404)
    try:
        if d.get("password"):
            pwd_err = validate_password_strength(d["password"])
            if pwd_err:
                db.close(); return error(pwd_err)
            db.execute(
                "UPDATE users SET name=?, username=?, email=?, password=?, role=?, active=? WHERE id=?",
                (d["name"].strip(), username, email, hash_password(d["password"]), role, int(d.get("active", user["active"])), user_id)
            )
        else:
            db.execute(
                "UPDATE users SET name=?, username=?, email=?, role=?, active=? WHERE id=?",
                (d["name"].strip(), username, email, role, int(d.get("active", user["active"])), user_id)
            )
        db.commit()
    except sqlite3.IntegrityError:
        db.close(); return error("Username or email already registered")
    db.close()
    audit(f"Updated user {username}", "Users")
    return success(msg="User updated")


@app.route("/api/users/<int:user_id>/role", methods=["PATCH"])
@roles_required("admin")
def assign_user_role(user_id):
    d = request.json or {}
    role = (d.get("role") or "").strip().lower()
    role_err = validate_user_role(role)
    if role_err:
        return error(role_err)
    if str(user_id) == str(g.user["sub"]) and role != "admin":
        return error("You cannot remove your own admin role")

    db = get_db()
    user = row_to_dict(db.execute("SELECT * FROM users WHERE id=?", (user_id,)).fetchone())
    if not user:
        db.close(); return error("User not found", 404)
    if user.get("role") == role:
        db.close()
        return success({
            "id": user_id,
            "role": role,
        }, "Role already assigned")

    db.execute("UPDATE users SET role=? WHERE id=?", (role, user_id))
    db.commit()
    updated = row_to_dict(db.execute(
        "SELECT id,name,username,email,role,active,created_at FROM users WHERE id=?",
        (user_id,),
    ).fetchone())
    db.close()
    audit(
        f"Assigned role {role} to {updated.get('username') or updated.get('email')}",
        "Users",
        f"Previous role: {user.get('role')}"
    )
    return success(updated, "Role assigned")


@app.route("/api/users/<int:user_id>/status", methods=["PATCH"])
@roles_required("admin")
def update_user_status(user_id):
    if str(user_id) == str(g.user["sub"]):
        return error("You cannot change your own account status")
    d = request.json or {}
    active = 1 if d.get("active") else 0
    db = get_db()
    user = row_to_dict(db.execute("SELECT * FROM users WHERE id=?", (user_id,)).fetchone())
    if not user:
        db.close(); return error("User not found", 404)
    db.execute("UPDATE users SET active=? WHERE id=?", (active, user_id))
    db.commit(); db.close()
    audit(f"{'Activated' if active else 'Deactivated'} user {user.get('username') or user.get('email')}", "Users")
    return success(msg="User activated" if active else "User deactivated")


@app.route("/api/users/<int:user_id>/password", methods=["PATCH"])
@roles_required("admin")
def reset_user_password(user_id):
    d = request.json or {}
    new_password = d.get("password", "")
    pwd_err = validate_password_strength(new_password)
    if pwd_err:
        return error(pwd_err)
    db = get_db()
    user = row_to_dict(db.execute("SELECT * FROM users WHERE id=?", (user_id,)).fetchone())
    if not user:
        db.close(); return error("User not found", 404)
    db.execute("UPDATE users SET password=? WHERE id=?", (hash_password(new_password), user_id))
    db.commit(); db.close()
    audit(f"Reset password for {user.get('username') or user.get('email')}", "Users")
    return success(msg="Password reset")


@app.route("/api/users/<int:user_id>", methods=["DELETE"])
@roles_required("admin")
def delete_user(user_id):
    if str(user_id) == str(g.user["sub"]):
        return error("You cannot delete your own account")
    db = get_db()
    user = row_to_dict(db.execute("SELECT * FROM users WHERE id=?", (user_id,)).fetchone())
    if not user:
        db.close(); return error("User not found", 404)
    db.execute("UPDATE members SET created_by=NULL WHERE created_by=?", (user_id,))
    db.execute("UPDATE loans SET approved_by=NULL WHERE approved_by=?", (user_id,))
    db.execute("UPDATE loans SET officer_id=NULL WHERE officer_id=?", (user_id,))
    db.execute("UPDATE repayments SET recorded_by=NULL WHERE recorded_by=?", (user_id,))
    db.execute("UPDATE savings_transactions SET recorded_by=NULL WHERE recorded_by=?", (user_id,))
    db.execute("DELETE FROM notifications WHERE user_id=?", (user_id,))
    db.execute("DELETE FROM audit_logs WHERE user_id=?", (user_id,))
    db.execute("DELETE FROM users WHERE id=?", (user_id,))
    db.commit(); db.close()
    audit(f"Deleted user {user.get('username') or user.get('email')}", "Users")
    return success(msg="User deleted")


@app.route("/api/settings", methods=["GET"])
def get_settings():
    return success(get_settings_dict())


@app.route("/api/settings", methods=["PUT"])
@roles_required("admin")
def update_settings():
    allowed = {"sacco_name", "logo_text", "logo_image", "logo_url", "address", "phone", "account_opening_balance"}
    d = request.json or {}
    db = get_db()
    for key, value in d.items():
        if key in allowed:
            db.execute(
                "INSERT INTO app_settings (key,value,updated_at) VALUES (?,?,datetime('now')) "
                "ON CONFLICT(key) DO UPDATE SET value=excluded.value, updated_at=datetime('now')",
                (key, str(value or ""))
            )
    db.commit(); db.close()
    audit("Updated application settings", "Settings")
    return success(get_settings_dict(), "Settings saved")


@app.route("/api/settings/account/add", methods=["POST"])
@roles_required("admin", "accountant")
def add_main_account_funds():
    d = request.json or {}
    try:
        amount = float(d.get("amount") or 0)
    except (TypeError, ValueError):
        return error("Amount must be a valid number")
    if amount <= 0:
        return error("Amount must be greater than zero")

    db = get_db()
    current_opening = get_account_opening_balance(db)
    new_opening = set_account_opening_balance(db, current_opening + amount)
    db.commit()
    db.close()

    audit("Added funds to main account", "Settings", f"KES {amount:,.2f}")
    settings = get_settings_dict()
    return success({
        "added_amount": amount,
        "account_opening_balance": float(settings.get("account_opening_balance") or 0),
        "settings": settings,
    }, "Main account updated")

# ══════════════════════════════════════════════════════════════════════════════
# LOAN CALCULATOR (public — no auth)
# ══════════════════════════════════════════════════════════════════════════════
@app.route("/api/calculate", methods=["POST"])
def calculate():
    d = request.json or {}
    try:
        principal   = float(d["amount"])
        annual_rate = float(d["annual_rate"])
        term_months = int(d["term_months"])
        method      = d.get("method","reducing")
        start       = d.get("start_date", date.today().isoformat())
    except (KeyError, ValueError, TypeError):
        return error("amount, annual_rate, term_months required")

    schedule = build_schedule("CALC", principal, annual_rate, term_months, method, start)
    total_repayable = sum(r["repayment"] for r in schedule)
    total_interest  = sum(r["interest"]  for r in schedule)
    return success({
        "monthly_repayment": schedule[0]["repayment"] if schedule else 0,
        "total_repayable":   round(total_repayable, 2),
        "total_interest":    round(total_interest, 2),
        "schedule":          schedule[:6],  # first 6 rows preview
    })


# ══════════════════════════════════════════════════════════════════════════════
# HEALTH CHECK
# ══════════════════════════════════════════════════════════════════════════════
@app.route("/api/health", methods=["GET"])
def health():
    return jsonify({"status":"ok","version":"3.0.0","timestamp": datetime.utcnow().isoformat()})


# ── Entry point ──────────────────────────────────────────────────────────────
if __name__ == "__main__":
    init_db()
    port = int(os.environ.get("PORT", "5000"))
    debug_mode = os.environ.get("DEBUG", "false").strip().lower() == "true"
    print(f"SACCOFinance API running on http://localhost:{port}")
    app.run(debug=debug_mode, port=port, host="0.0.0.0")

