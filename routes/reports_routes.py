from collections import Counter, defaultdict
from datetime import datetime
from io import BytesIO

from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.styles import Alignment, Font, PatternFill

from api import Blueprint

from services.common import *

reports_bp = Blueprint("reports", __name__)
bp = reports_bp


def _xlsx_style_header(ws, row=1):
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in ws[row]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")


def _xlsx_autosize(ws):
    for col in ws.columns:
        cells = list(col)
        if not cells:
            continue
        max_len = max((len(str(c.value or "")) for c in cells), default=10)
        ws.column_dimensions[cells[0].column_letter].width = min(42, max(12, max_len + 2))


def _xlsx_cell_value(row, header):
    if not isinstance(row, dict):
        return row
    variants = [
        header,
        header.lower(),
        header.upper(),
        header.replace(" ", "_"),
        header.replace(" ", "_").lower(),
        header.replace(" ", "_").title(),
    ]
    for key in variants:
        if key in row and row[key] is not None:
            return row[key]
    return None

def _xlsx_write_rows(ws, headers, rows):
    ws.append(headers)
    for row in rows:
        if isinstance(row, dict):
            ws.append([_xlsx_cell_value(row, h) for h in headers])
        else:
            ws.append(list(row))
    _xlsx_style_header(ws)
    _xlsx_autosize(ws)


def _xlsx_summary_sheet(wb, report_type: str, headers: list[str], rows: list[dict], metrics: dict) -> None:
    ws = wb.create_sheet("Summary")
    ws["A1"] = f"{report_type.replace('-', ' ').title()} Report"
    ws["A1"].font = Font(bold=True, size=16)
    ws["A3"] = "Generated At"
    ws["B3"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ws["A4"] = "Record Count"
    ws["B4"] = len(rows)
    ws["A6"] = "Key Metrics"
    ws["A6"].font = Font(bold=True, size=12)
    ws["A7"] = "Metric"
    ws["B7"] = "Value"
    for idx, (key, value) in enumerate(metrics.items(), start=8):
        ws[f"A{idx}"] = key
        ws[f"B{idx}"] = value
    _xlsx_style_header(ws, 7)
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 20


def _xlsx_analytics_sheet(wb, report_type: str, rows: list[dict]) -> None:
    ws = wb.create_sheet("Analytics")
    ws["A1"] = "Analytics"
    ws["A1"].font = Font(bold=True, size=14)

    if not rows:
        ws["A3"] = "No records available for analytics."
        return

    if report_type == "loans":
        status_counts = Counter(str(row.get("Status") or row.get("status") or "").title() or "Unknown" for row in rows)
        ws["A3"] = "Loan Status"
        ws["B3"] = "Count"
        for idx, (status, count) in enumerate(status_counts.items(), start=4):
            ws[f"A{idx}"] = status
            ws[f"B{idx}"] = count
        _xlsx_style_header(ws, 3)
        pie = PieChart()
        pie.title = "Loan Status Breakdown"
        data = Reference(ws, min_col=2, min_row=3, max_row=3 + len(status_counts))
        labels = Reference(ws, min_col=1, min_row=4, max_row=3 + len(status_counts))
        pie.add_data(data, titles_from_data=True)
        pie.set_categories(labels)
        pie.height = 7
        pie.width = 9
        pie.dataLabels = DataLabelList()
        pie.dataLabels.showPercent = True
        ws.add_chart(pie, "D3")

        top = sorted(rows, key=lambda row: float(row.get("Outstanding") or row.get("outstanding") or 0), reverse=True)[:10]
        start = 20
        ws[f"A{start}"] = "Top Outstanding Loans"
        ws[f"B{start}"] = "Outstanding"
        for idx, row in enumerate(top, start=start + 1):
            ws[f"A{idx}"] = row.get("ID") or row.get("id")
            ws[f"B{idx}"] = float(row.get("Outstanding") or row.get("outstanding") or 0)
        _xlsx_style_header(ws, start)
        bar = BarChart()
        bar.type = "bar"
        bar.style = 10
        bar.title = "Top Outstanding Loans"
        bar.y_axis.title = "Loan"
        bar.x_axis.title = "KES"
        data = Reference(ws, min_col=2, min_row=start, max_row=start + len(top))
        cats = Reference(ws, min_col=1, min_row=start + 1, max_row=start + len(top))
        bar.add_data(data, titles_from_data=True)
        bar.set_categories(cats)
        bar.height = 7
        bar.width = 13
        ws.add_chart(bar, "D20")

    elif report_type == "account-monthly":
        ws["A3"] = "Month"
        ws["B3"] = "Opening"
        ws["C3"] = "Inflow"
        ws["D3"] = "Outflow"
        ws["E3"] = "Closing"
        for idx, row in enumerate(rows, start=4):
            ws[f"A{idx}"] = row.get("Month") or row.get("month")
            ws[f"B{idx}"] = float(row.get("Opening Balance") or row.get("opening_balance") or 0)
            ws[f"C{idx}"] = float(row.get("Total Inflow") or row.get("inflow") or 0)
            ws[f"D{idx}"] = float(row.get("Total Outflow") or row.get("outflow") or 0)
            ws[f"E{idx}"] = float(row.get("Closing Balance") or row.get("closing_balance") or 0)
        _xlsx_style_header(ws, 3)
        line = LineChart()
        line.title = "Monthly Cashflow"
        line.y_axis.title = "KES"
        line.x_axis.title = "Month"
        data = Reference(ws, min_col=2, max_col=5, min_row=3, max_row=3 + len(rows))
        cats = Reference(ws, min_col=1, min_row=4, max_row=3 + len(rows))
        line.add_data(data, titles_from_data=True)
        line.set_categories(cats)
        line.height = 7
        line.width = 14
        ws.add_chart(line, "G3")

    elif report_type == "savings":
        balances = sorted(rows, key=lambda row: float(row.get("Balance") or row.get("balance") or 0), reverse=True)[:10]
        ws["A3"] = "Top Savers"
        ws["B3"] = "Balance"
        for idx, row in enumerate(balances, start=4):
            ws[f"A{idx}"] = row.get("Name") or row.get("name")
            ws[f"B{idx}"] = float(row.get("Balance") or row.get("balance") or 0)
        _xlsx_style_header(ws, 3)
        bar = BarChart()
        bar.type = "bar"
        bar.title = "Top Savers"
        bar.y_axis.title = "Member"
        bar.x_axis.title = "KES"
        data = Reference(ws, min_col=2, min_row=3, max_row=3 + len(balances))
        cats = Reference(ws, min_col=1, min_row=4, max_row=3 + len(balances))
        bar.add_data(data, titles_from_data=True)
        bar.set_categories(cats)
        bar.height = 7
        bar.width = 13
        ws.add_chart(bar, "D3")

    elif report_type == "expenses":
        grouped = defaultdict(float)
        for row in rows:
            grouped[str(row.get("Account") or row.get("account") or "Unknown")] += float(row.get("Amount") or row.get("amount") or 0)
        ws["A3"] = "Account"
        ws["B3"] = "Total Expense"
        for idx, (name, amount) in enumerate(sorted(grouped.items(), key=lambda item: item[1], reverse=True), start=4):
            ws[f"A{idx}"] = name
            ws[f"B{idx}"] = amount
        _xlsx_style_header(ws, 3)
        bar = BarChart()
        bar.type = "bar"
        bar.title = "Expense by Account"
        bar.x_axis.title = "KES"
        data = Reference(ws, min_col=2, min_row=3, max_row=3 + len(grouped))
        cats = Reference(ws, min_col=1, min_row=4, max_row=3 + len(grouped))
        bar.add_data(data, titles_from_data=True)
        bar.set_categories(cats)
        bar.height = 7
        bar.width = 13
        ws.add_chart(bar, "D3")

    elif report_type == "repayments":
        monthly = defaultdict(float)
        for row in rows:
            payment_date = str(row.get("Date") or row.get("payment_date") or "")[:7]
            if payment_date:
                monthly[payment_date] += float(row.get("Amount") or row.get("amount") or 0)
        ws["A3"] = "Month"
        ws["B3"] = "Repayments"
        for idx, (month, amount) in enumerate(sorted(monthly.items()), start=4):
            ws[f"A{idx}"] = month
            ws[f"B{idx}"] = amount
        _xlsx_style_header(ws, 3)
        line = LineChart()
        line.title = "Monthly Repayments"
        line.x_axis.title = "Month"
        line.y_axis.title = "KES"
        data = Reference(ws, min_col=2, min_row=3, max_row=3 + len(monthly))
        cats = Reference(ws, min_col=1, min_row=4, max_row=3 + len(monthly))
        line.add_data(data, titles_from_data=True)
        line.set_categories(cats)
        line.height = 7
        line.width = 14
        ws.add_chart(line, "D3")

    elif report_type == "dividends":
        ranked = sorted(rows, key=lambda row: float(row.get("Dividend Amount") or row.get("dividend_amount") or 0), reverse=True)[:10]
        ws["A3"] = "Member"
        ws["B3"] = "Dividend"
        for idx, row in enumerate(ranked, start=4):
            ws[f"A{idx}"] = row.get("Member") or row.get("member")
            ws[f"B{idx}"] = float(row.get("Dividend Amount") or row.get("dividend_amount") or 0)
        _xlsx_style_header(ws, 3)
        bar = BarChart()
        bar.type = "bar"
        bar.title = "Top Dividend Allocations"
        data = Reference(ws, min_col=2, min_row=3, max_row=3 + len(ranked))
        cats = Reference(ws, min_col=1, min_row=4, max_row=3 + len(ranked))
        bar.add_data(data, titles_from_data=True)
        bar.set_categories(cats)
        bar.height = 7
        bar.width = 13
        ws.add_chart(bar, "D3")

    else:
        ws["A3"] = "Analytics"
        ws["A4"] = "No report-specific analytics available."


def _xlsx_metrics(report_type: str, rows: list[dict], totals: dict | None = None) -> dict:
    totals = totals or {}
    if report_type == "loans":
        return {
            "Total Loans": len(rows),
            "Total Disbursed": round(sum(float(row.get("Amount") or row.get("amount") or 0) for row in rows), 2),
            "Total Repaid": round(sum(float(row.get("Paid") or row.get("total_paid") or 0) for row in rows), 2),
            "Total Outstanding": round(sum(float(row.get("Outstanding") or row.get("outstanding") or 0) for row in rows), 2),
            "Total Penalties": round(sum(float(row.get("Penalties") or row.get("penalties") or 0) for row in rows), 2),
        }
    if report_type == "account-monthly":
        last = rows[-1] if rows else {}
        return {
            "Months": len(rows),
            "Closing Balance": round(float(last.get("Closing Balance") or last.get("closing_balance") or 0), 2) if last else 0,
            "Total Inflow": round(sum(float(row.get("Total Inflow") or row.get("inflow") or 0) for row in rows), 2),
            "Total Outflow": round(sum(float(row.get("Total Outflow") or row.get("outflow") or 0) for row in rows), 2),
        }
    if report_type == "savings":
        return {
            "Members": len(rows),
            "Total Savings": round(sum(float(row.get("Balance") or row.get("balance") or 0) for row in rows), 2),
        }
    if report_type == "repayments":
        return {
            "Repayments": len(rows),
            "Total Repaid": round(sum(float(row.get("Amount") or row.get("amount") or 0) for row in rows), 2),
        }
    if report_type == "expenses":
        return {
            "Transactions": len(rows),
            "Total Expenses": round(sum(float(row.get("Amount") or row.get("amount") or 0) for row in rows), 2),
        }
    if report_type == "dividends":
        return {
            "Allocations": len(rows),
            "Total Dividend": round(sum(float(row.get("Dividend Amount") or row.get("dividend_amount") or 0) for row in rows), 2),
        }
    return {"Rows": len(rows)}

@bp.route("/api/reports/portfolio", methods=["GET"])
@login_required
def report_portfolio():
    db = get_db()
    refresh_loan_statuses(db)
    db.commit()
    month_filter = (request.args.get("month") or "").strip()
    month_clause = " AND strftime('%Y-%m', l.disbursed_date)=?" if month_filter else ""
    month_params = (month_filter,) if month_filter else ()
    available_months = rows_to_list(db.execute(
        """SELECT DISTINCT strftime('%Y-%m', disbursed_date) AS month
           FROM loans
           WHERE disbursed_date IS NOT NULL
           ORDER BY month DESC"""
    ).fetchall())
    rows = rows_to_list(db.execute(
        f"""SELECT l.*, m.name as member_name,
           COALESCE(risk.total_repayable, l.amount) + COALESCE(l.penalties,0) as total_repayable,
           CASE WHEN l.status='written_off' THEN 0 ELSE MAX(COALESCE(risk.total_repayable, l.amount) + COALESCE(l.penalties,0) - l.total_paid, 0) END as outstanding,
           CASE WHEN l.status='written_off' THEN 0 ELSE COALESCE(risk.amount_in_arrears,0) END as amount_in_arrears,
           COALESCE(risk.overdue_installments,0) as overdue_installments,
           COALESCE(CAST(julianday(date('now')) - julianday(risk.oldest_due_date) AS INTEGER),0) as days_in_arrears,
           risk.next_due_date
           FROM loans l
           JOIN members m ON l.member_id=m.id
           LEFT JOIN (
             SELECT loan_id,
                    SUM(repayment) as total_repayable,
                    MIN(CASE WHEN paid=0 THEN due_date END) as next_due_date,
                    MIN(CASE WHEN paid=0 AND due_date < date('now') THEN due_date END) as oldest_due_date,
                    SUM(CASE WHEN paid=0 AND due_date < date('now') THEN repayment + COALESCE(penalty,0) ELSE 0 END) as amount_in_arrears,
                    COUNT(CASE WHEN paid=0 AND due_date < date('now') THEN 1 END) as overdue_installments
             FROM loan_schedule
             GROUP BY loan_id
           ) risk ON risk.loan_id=l.id
           WHERE l.disbursed_date IS NOT NULL{month_clause}
           ORDER BY l.disbursed_date DESC""",
        month_params
    ).fetchall())
    totals = db.execute(
        f"""SELECT
           COALESCE(SUM(amount),0) as total_disbursed,
           COALESCE(SUM(total_paid),0) as total_repaid,
           COALESCE(SUM(CASE WHEN status='written_off' THEN 0 ELSE total_repayable END),0) as total_repayable,
           COALESCE(SUM(CASE WHEN status='written_off' THEN 0 ELSE MAX(total_repayable - total_paid, 0) END),0) as total_outstanding,
           COALESCE(SUM(CASE WHEN status='written_off' THEN 0 ELSE amount_in_arrears END),0) as amount_in_arrears,
           COALESCE(SUM(CASE WHEN status='written_off' THEN 0 WHEN days_in_arrears >= 30 THEN MAX(total_repayable - total_paid, 0) ELSE 0 END),0) as par30_amount,
           COALESCE(SUM(penalties),0) as total_penalties,
           COUNT(*) as total_loans
           FROM (
                SELECT l.id, l.amount, l.total_paid, l.penalties, l.disbursed_date, l.status,
                        COALESCE(risk.total_repayable, l.amount) + COALESCE(l.penalties,0) as total_repayable,
                        COALESCE(risk.amount_in_arrears,0) as amount_in_arrears,
                        risk.days_in_arrears
                FROM loans l
                LEFT JOIN (
                  SELECT loan_id,
                         SUM(repayment) as total_repayable,
                         SUM(CASE WHEN paid=0 AND due_date < date('now') THEN repayment + COALESCE(penalty,0) ELSE 0 END) as amount_in_arrears,
                         CAST(julianday(date('now')) - julianday(MIN(CASE WHEN paid=0 AND due_date < date('now') THEN due_date END)) AS INTEGER) as days_in_arrears
                  FROM loan_schedule
                  GROUP BY loan_id
                ) risk ON risk.loan_id=l.id
                WHERE l.disbursed_date IS NOT NULL{month_clause}
            )""",
        month_params
    ).fetchone()
    regions = rows_to_list(db.execute(
        f"""SELECT LOWER(COALESCE(NULLIF(TRIM(m.region), ''), 'unknown')) AS region_key,
                  COUNT(*) AS loan_count,
                  COALESCE(SUM(l.amount),0) AS total_disbursed,
                  COALESCE(SUM(CASE WHEN l.status='written_off' THEN 0 ELSE MAX(COALESCE(risk.total_repayable, l.amount) + COALESCE(l.penalties,0) - l.total_paid, 0) END),0) AS outstanding
           FROM loans l
           JOIN members m ON m.id=l.member_id
           LEFT JOIN (
             SELECT loan_id,
                    SUM(repayment) as total_repayable
             FROM loan_schedule
             GROUP BY loan_id
           ) risk ON risk.loan_id=l.id
           WHERE l.disbursed_date IS NOT NULL{month_clause}
           GROUP BY region_key
           ORDER BY total_disbursed DESC, loan_count DESC""",
        month_params
    ).fetchall())
    for row in regions:
        raw_region = (row.get("region_key") or "unknown").strip()
        row["region"] = raw_region.title() if raw_region else "Unknown"
        row.pop("region_key", None)
    db.close()
    return success({"loans": rows, "totals": dict(totals), "months": available_months, "selected_month": month_filter, "regions": regions})



@bp.route("/api/reports/savings", methods=["GET"])
@login_required
def report_savings():
    db = get_db()
    rows = rows_to_list(db.execute(
        """SELECT m.id, m.name, m.status, sa.balance,
           COALESCE(SUM(CASE WHEN st.type='deposit' THEN st.amount ELSE 0 END),0) as total_deposits,
           COALESCE(SUM(CASE WHEN st.type='withdrawal' THEN st.amount ELSE 0 END),0) as total_withdrawals
           FROM members m
           LEFT JOIN savings_accounts sa ON m.id=sa.member_id
           LEFT JOIN savings_transactions st ON m.id=st.member_id
           WHERE m.member_type='member'
           GROUP BY m.id ORDER BY sa.balance DESC"""
    ).fetchall())
    db.close()
    return success(rows)



@bp.route("/api/reports/account-monthly", methods=["GET"])
@login_required
def report_account_monthly():
    db = get_db()
    data = build_monthly_account_report(db)
    db.close()
    return success(data)



@bp.route("/api/dividends", methods=["GET"])
@login_required
def list_dividends():
    year = request.args.get("year")
    where = []
    params = []
    if year:
        where.append("dr.year=?")
        params.append(int(year))
    clause = "WHERE " + " AND ".join(where) if where else ""
    db = get_db()
    runs = rows_to_list(db.execute(
        f"""SELECT dr.*, u.name AS created_by_name,
                  COUNT(da.id) AS member_count,
                  COALESCE(SUM(da.dividend_amount),0) AS allocated_total
           FROM dividend_runs dr
           LEFT JOIN dividend_allocations da ON da.run_id=dr.id
           LEFT JOIN users u ON u.id=dr.created_by
           {clause}
           GROUP BY dr.id
           ORDER BY dr.year DESC, dr.created_at DESC""",
        params,
    ).fetchall())
    selected = None
    allocations = []
    if runs:
        selected = runs[0]
        allocations = rows_to_list(db.execute(
            """SELECT da.*, m.name AS member_name, m.phone AS member_phone
               FROM dividend_allocations da JOIN members m ON m.id=da.member_id
               WHERE da.run_id=? ORDER BY da.dividend_amount DESC""",
            (selected["id"],),
        ).fetchall())
    db.close()
    return success({"runs": runs, "selected": selected, "allocations": allocations})


@bp.route("/api/dividends/calculate", methods=["POST"])
@login_required
@roles_required("admin", "accountant")
def calculate_dividends():
    d = request.json or {}
    try:
        year = int(d.get("year") or date.today().year)
        surplus = float(d.get("surplus") or 0)
    except (TypeError, ValueError):
        return error("Year and surplus must be valid numbers")
    if surplus <= 0:
        return error("Surplus must be greater than zero")
    basis = (d.get("basis") or "savings_balance").strip().lower()
    if basis != "savings_balance":
        return error("Only savings_balance dividend basis is currently supported")
    db = get_db()
    members = rows_to_list(db.execute(
        """SELECT m.id, m.name, COALESCE(sa.balance, m.savings, 0) AS basis_amount
           FROM members m LEFT JOIN savings_accounts sa ON sa.member_id=m.id
           WHERE m.member_type='member' AND m.status='active'
           ORDER BY m.name"""
    ).fetchall())
    total_basis = sum(float(member.get("basis_amount") or 0) for member in members)
    if total_basis <= 0:
        db.close(); return error("No member savings balance available for dividend allocation")
    db.execute(
        "INSERT INTO dividend_runs (year,surplus,basis,total_basis,status,created_by) VALUES (?,?,?,?,?,?)",
        (year, surplus, basis, total_basis, "draft", g.user["sub"]),
    )
    run_id = db.execute("SELECT last_insert_rowid()").fetchone()[0]
    allocations = []
    remaining = round(surplus, 2)
    eligible = [member for member in members if float(member.get("basis_amount") or 0) > 0]
    for index, member in enumerate(eligible):
        basis_amount = float(member.get("basis_amount") or 0)
        amount = remaining if index == len(eligible) - 1 else round((basis_amount / total_basis) * surplus, 2)
        remaining = round(remaining - amount, 2)
        db.execute(
            "INSERT INTO dividend_allocations (run_id,member_id,basis_amount,dividend_amount) VALUES (?,?,?,?)",
            (run_id, member["id"], basis_amount, amount),
        )
        allocations.append({"member_id": member["id"], "member_name": member["name"], "basis_amount": basis_amount, "dividend_amount": amount})
    db.commit()
    run = row_to_dict(db.execute("SELECT * FROM dividend_runs WHERE id=?", (run_id,)).fetchone())
    db.close()
    audit(f"Calculated dividends for {year}", "Dividends", f"Surplus KES {surplus}")
    return success({"run": run, "allocations": allocations}, "Dividend allocation calculated", 201)


@bp.route("/api/dividends/<int:run_id>/status", methods=["PATCH"])
@login_required
@roles_required("admin", "accountant")
def update_dividend_status(run_id):
    status = ((request.json or {}).get("status") or "").strip().lower()
    if status not in {"draft", "approved", "paid", "cancelled"}:
        return error("Status must be draft, approved, paid, or cancelled")
    db = get_db()
    run = row_to_dict(db.execute("SELECT * FROM dividend_runs WHERE id=?", (run_id,)).fetchone())
    if not run:
        db.close(); return error("Dividend run not found", 404)
    db.execute("UPDATE dividend_runs SET status=? WHERE id=?", (status, run_id))
    if status == "paid":
        db.execute("UPDATE dividend_allocations SET paid=1 WHERE run_id=?", (run_id,))
    db.commit(); db.close()
    audit(f"Set dividend run {run_id} to {status}", "Dividends")
    return success(msg="Dividend status updated")


@bp.route("/api/reports/export/<report_type>", methods=["GET"])
@login_required
def export_report(report_type):
    import csv, io
    db = get_db()
    refresh_loan_statuses(db)
    db.commit()
    export_format = (request.args.get("format") or "csv").strip().lower()
    month_filter = (request.args.get("month") or "").strip()
    month_clause = " AND strftime('%Y-%m', l.disbursed_date)=?" if month_filter else ""
    month_params = (month_filter,) if month_filter else ()
    rows = []
    headers = []

    if report_type == "loans":
        headers = ["ID","Borrower Name","Member Type","Amount","Rate%","Term","Method","Status","Disbursed","Paid","Outstanding","Penalties"]
        rows = rows_to_list(db.execute(
            f"""SELECT l.id,m.name AS 'Borrower Name',m.member_type AS 'Member Type',l.amount,l.annual_rate,l.term_months,l.method,l.status,l.disbursed_date,l.total_paid,
                      CASE WHEN l.status='written_off' THEN 0 ELSE MAX(COALESCE(SUM(s.repayment), l.amount)+COALESCE(l.penalties,0)-l.total_paid, 0) END as outstanding,l.penalties
                FROM loans l
                JOIN members m ON l.member_id=m.id
                LEFT JOIN loan_schedule s ON s.loan_id=l.id
                WHERE l.disbursed_date IS NOT NULL{month_clause}
                GROUP BY l.id""",
            month_params
        ).fetchall())
    elif report_type == "repayments":
        headers = ["ID","Loan","Borrower Name","Amount","Date","Method","Reference","Type"]
        rows = rows_to_list(db.execute(
            "SELECT r.id,r.loan_id,m.name AS 'Borrower Name',r.amount,r.payment_date,r.method,r.reference,r.type FROM repayments r JOIN members m ON r.member_id=m.id ORDER BY r.payment_date DESC"
        ).fetchall())
    elif report_type == "savings":
        headers = ["Member ID","Member","Balance","Status"]
        rows = rows_to_list(db.execute(
            "SELECT m.id,m.name,sa.balance,m.status FROM members m LEFT JOIN savings_accounts sa ON m.id=sa.member_id WHERE m.member_type='member' ORDER BY sa.balance DESC"
        ).fetchall())
    elif report_type == "expenses":
        headers = ["ID","Date","Account","Account Code","Amount","Payee","Reference","Notes","Recorded By"]
        rows = rows_to_list(db.execute(
            """SELECT et.id, et.expense_date, ea.name, ea.code, et.amount, et.payee, et.reference, et.notes, u.name
               FROM expense_transactions et
               JOIN expense_accounts ea ON et.account_id=ea.id
               LEFT JOIN users u ON et.recorded_by=u.id
               ORDER BY et.expense_date DESC, et.created_at DESC"""
        ).fetchall())
    elif report_type == "account-monthly":
        headers = ["Month","Opening Balance","Savings Collections","Loan Repayments","Total Inflow","Loans Disbursed","Expenses","Total Outflow","Net Movement","Closing Balance"]
        data = build_monthly_account_report(db)
        for row in data.get("months", []):
            rows.append({
                "Month": row.get("month"),
                "Opening Balance": row.get("opening_balance"),
                "Savings Collections": row.get("savings_collections"),
                "Loan Repayments": row.get("loan_repayments"),
                "Total Inflow": row.get("inflow"),
                "Loans Disbursed": row.get("loan_disbursed"),
                "Expenses": row.get("expenses"),
                "Total Outflow": row.get("outflow"),
                "Net Movement": row.get("net"),
                "Closing Balance": row.get("closing_balance"),
            })
    elif report_type == "members":
        headers = ["ID","Name","Phone","Email","National ID","Status","Joined","Savings"]
        rows = rows_to_list(db.execute("SELECT id,name,phone,email,national_id,status,joined_date,savings FROM members WHERE member_type='member'").fetchall())
    elif report_type == "dividends":
        headers = ["Run ID","Year","Member ID","Member","Basis Amount","Dividend Amount","Paid","Status"]
        rows = rows_to_list(db.execute(
            """SELECT dr.id AS 'Run ID', dr.year AS Year, da.member_id AS 'Member ID', m.name AS Member,
                      da.basis_amount AS 'Basis Amount', da.dividend_amount AS 'Dividend Amount', da.paid AS Paid, dr.status AS Status
               FROM dividend_allocations da
               JOIN dividend_runs dr ON dr.id=da.run_id
               JOIN members m ON m.id=da.member_id
               ORDER BY dr.year DESC, da.dividend_amount DESC"""
        ).fetchall())
    else:
        db.close(); return error("Unknown report type")

    # CSV output is the default and remains backward-compatible.
    if export_format not in ("csv", "xlsx"):
        db.close()
        return error("Unsupported export format. Use csv or xlsx.")

    db.close()
    from api import Response
    audit(f"Exported {report_type} report", "Reports")
    if export_format == "xlsx":
        try:
            from openpyxl import Workbook
        except Exception:
            return error("Excel export dependency missing. Install openpyxl.")

        wb = Workbook()
        wb.remove(wb.active)
        data_rows = []
        for row in rows:
            if isinstance(row, dict):
                data_rows.append(row)
            else:
                data_rows.append(dict(zip(headers, row)))
        metrics = _xlsx_metrics(report_type, data_rows)
        _xlsx_summary_sheet(wb, report_type, headers, data_rows, metrics)
        data_ws = wb.create_sheet("Data")
        _xlsx_write_rows(data_ws, headers, data_rows)
        _xlsx_analytics_sheet(wb, report_type, data_rows)

        stream = BytesIO()
        wb.save(stream)
        stream.seek(0)
        return send_file(
            stream,
            as_attachment=True,
            download_name=f"{report_type}-report{('-' + month_filter) if month_filter and report_type == 'loans' else ''}.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(headers)
    for row in rows:
        writer.writerow([row.get(h) for h in headers] if isinstance(row, dict) else list(row))
    return Response(
        output.getvalue(),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment;filename={report_type}-report{('-' + month_filter) if month_filter and report_type == 'loans' else ''}.csv"}
    )


# ══════════════════════════════════════════════════════════════════════════════
# ACCOUNTING
# ══════════════════════════════════════════════════════════════════════════════
