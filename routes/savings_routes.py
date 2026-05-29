from flask import Blueprint

from services.common import *

savings_bp = Blueprint("savings", __name__)
bp = savings_bp

@bp.route("/api/savings", methods=["GET"])
@login_required
def get_savings():
    db   = get_db()
    rows = rows_to_list(db.execute(
        """SELECT m.id, m.name, m.phone, m.status, sa.balance,
           COUNT(st.id) as txn_count
           FROM members m
           LEFT JOIN savings_accounts sa ON m.id=sa.member_id
           LEFT JOIN savings_transactions st ON m.id=st.member_id
           WHERE m.member_type='member'
           GROUP BY m.id ORDER BY sa.balance DESC"""
    ).fetchall())
    db.close()
    return success(rows)



@bp.route("/api/savings/transactions", methods=["GET"])
@login_required
def get_savings_transactions():
    member_id = request.args.get("member_id","")
    q         = request.args.get("q","")
    page      = max(1, int(request.args.get("page",1)))
    limit     = int(request.args.get("limit",15))

    where  = []
    params = []
    if member_id:
        where.append("st.member_id=?"); params.append(member_id)
    if q:
        where.append("(m.name LIKE ? OR st.reference LIKE ?)"); params.extend([f"%{q}%"]*2)

    clause = ("WHERE " + " AND ".join(where)) if where else ""
    db     = get_db()
    total  = db.execute(
        f"SELECT COUNT(*) FROM savings_transactions st JOIN members m ON st.member_id=m.id {clause}", params
    ).fetchone()[0]
    rows   = rows_to_list(db.execute(
        f"""SELECT st.*, m.name as member_name
            FROM savings_transactions st JOIN members m ON st.member_id=m.id
            {clause} ORDER BY st.txn_date DESC, st.created_at DESC LIMIT ? OFFSET ?""",
        params + [limit, (page-1)*limit]
    ).fetchall())
    db.close()
    return success({"transactions": rows, "total": total, "page": page, "limit": limit, "pages": -(-total//limit)})



@bp.route("/api/savings/deposit", methods=["POST"])
@login_required
@roles_required("admin","officer","cashier","accountant")
def savings_deposit():
    d = request.json or {}
    if not d.get("member_id") or not d.get("amount"):
        return error("member_id and amount required")

    amount = float(d["amount"])
    if amount <= 0:
        return error("Amount must be positive")

    db     = get_db()
    member = row_to_dict(db.execute("SELECT * FROM members WHERE id=?", (d["member_id"],)).fetchone())
    if not member:
        db.close(); return error("Member not found")

    new_balance = member["savings"] + amount
    tid = gen_id("ST")
    ref = d.get("reference") or "DEP" + gen_id()

    db.execute("UPDATE members SET savings=? WHERE id=?", (new_balance, d["member_id"]))
    db.execute("UPDATE savings_accounts SET balance=? WHERE member_id=?", (new_balance, d["member_id"]))
    db.execute(
        "INSERT INTO savings_transactions (id,member_id,type,amount,category,txn_date,reference,balance_after,recorded_by) VALUES (?,?,?,?,?,?,?,?,?)",
        (tid, d["member_id"], "deposit", amount, d.get("category","voluntary"),
         d.get("date", date.today().isoformat()), ref, new_balance, g.user["sub"])
    )
    adjust_account_opening_balance(db, amount)
    db.commit()
    db.close()
    audit(f"Recorded deposit {tid}", "Savings", f"KES {amount} for {d['member_id']}")
    return success({"reference": ref, "new_balance": new_balance}, "Deposit recorded", 201)



@bp.route("/api/savings/bulk-deposit", methods=["POST"])
@login_required
@roles_required("admin", "officer", "cashier", "accountant")
def savings_bulk_deposit():
    d = request.json or {}
    items = d.get("items") or []
    if not isinstance(items, list) or not items:
        return error("At least one deposit row is required")
    txn_date = clean_date(d.get("date"), date.today().isoformat())
    category = (d.get("category") or "mandatory").strip() or "mandatory"
    batch_ref = (d.get("reference") or f"BULK-{gen_id()}").strip()
    db = get_db()
    created = []
    total = 0.0
    try:
        for idx, item in enumerate(items, start=1):
            member_id = (item.get("member_id") or "").strip()
            amount = float(item.get("amount") or 0)
            if not member_id or amount <= 0:
                continue
            member = row_to_dict(db.execute("SELECT * FROM members WHERE id=? AND member_type='member'", (member_id,)).fetchone())
            if not member:
                raise ValueError(f"Member not found: {member_id}")
            new_balance = float(member.get("savings") or 0) + amount
            tid = gen_id("ST")
            ref = item.get("reference") or f"{batch_ref}-{idx:02d}"
            db.execute("UPDATE members SET savings=? WHERE id=?", (new_balance, member_id))
            db.execute("UPDATE savings_accounts SET balance=? WHERE member_id=?", (new_balance, member_id))
            db.execute(
                "INSERT INTO savings_transactions (id,member_id,type,amount,category,txn_date,reference,balance_after,recorded_by) VALUES (?,?,?,?,?,?,?,?,?)",
                (tid, member_id, "deposit", amount, category, txn_date, ref, new_balance, g.user["sub"]),
            )
            created.append({"id": tid, "member_id": member_id, "amount": amount, "reference": ref, "balance_after": new_balance})
            total += amount
        if not created:
            db.close(); return error("No valid deposit rows were submitted")
        adjust_account_opening_balance(db, total)
        db.commit()
    except ValueError as exc:
        db.rollback(); db.close(); return error(str(exc))
    except sqlite3.IntegrityError as exc:
        db.rollback(); db.close(); return error(f"Bulk deposit failed: {exc}")
    db.close()
    audit("Recorded bulk savings deposit", "Savings", f"{len(created)} deposits; KES {total}")
    return success({"count": len(created), "total": total, "transactions": created, "reference": batch_ref}, "Bulk deposit recorded", 201)


def _passbook_rows(db, member_id):
    rows = rows_to_list(db.execute(
        """SELECT st.*, m.name AS member_name, m.phone AS member_phone
           FROM savings_transactions st JOIN members m ON m.id=st.member_id
           WHERE st.member_id=?
           ORDER BY st.txn_date ASC, st.created_at ASC, st.id ASC""",
        (member_id,),
    ).fetchall())
    running = 0.0
    for row in rows:
        amount = float(row.get("amount") or 0)
        if row.get("type") == "withdrawal":
            running -= amount
        else:
            running += amount
        row["running_balance"] = round(running, 2)
    return rows


@bp.route("/api/members/<member_id>/passbook", methods=["GET"])
@login_required
def member_passbook(member_id):
    export_format = (request.args.get("format") or "json").strip().lower()
    db = get_db()
    member = row_to_dict(db.execute("SELECT * FROM members WHERE id=?", (member_id,)).fetchone())
    if not member:
        db.close(); return error("Member not found", 404)
    rows = _passbook_rows(db, member_id)
    db.close()

    if export_format == "json":
        return success({"member": member, "transactions": rows})

    if export_format == "pdf":
        lines = [
            f"Member Passbook - {member.get('name')} ({member_id})",
            f"Generated: {date.today().isoformat()}",
            "",
            "Date | Type | Amount | Reference | Balance",
        ]
        for row in rows:
            sign_amount = -float(row.get("amount") or 0) if row.get("type") == "withdrawal" else float(row.get("amount") or 0)
            lines.append(f"{row.get('txn_date')} | {row.get('type')} | KES {sign_amount:,.2f} | {row.get('reference') or '-'} | KES {row.get('running_balance', 0):,.2f}")
        pdf = make_simple_pdf(lines[:48], title=f"Passbook {member_id}")
        return send_file(BytesIO(pdf), mimetype="application/pdf", as_attachment=True, download_name=f"passbook-{member_id}.pdf")

    if export_format == "xlsx":
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill
        except Exception:
            return error("Excel export dependency missing. Install openpyxl.")
        wb = Workbook()
        ws = wb.active
        ws.title = "Passbook"
        ws.append(["Date", "Type", "Amount", "Category", "Reference", "Balance"])
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="E5EEF9", end_color="E5EEF9", fill_type="solid")
        for row in rows:
            amount = -float(row.get("amount") or 0) if row.get("type") == "withdrawal" else float(row.get("amount") or 0)
            ws.append([row.get("txn_date"), row.get("type"), amount, row.get("category"), row.get("reference"), row.get("running_balance")])
        for col in ws.columns:
            max_len = max((len(str(c.value or "")) for c in col), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(42, max(12, max_len + 2))
        stream = BytesIO()
        wb.save(stream)
        stream.seek(0)
        return send_file(stream, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", as_attachment=True, download_name=f"passbook-{member_id}.xlsx")

    return error("Unsupported passbook format. Use json, pdf, or xlsx.")


@bp.route("/api/savings/withdraw", methods=["POST"])
@login_required
@roles_required("admin","officer","cashier","accountant")
def savings_withdraw():
    d = request.json or {}
    if not d.get("member_id") or not d.get("amount"):
        return error("member_id and amount required")

    amount = float(d["amount"])
    if amount <= 0:
        return error("Amount must be positive")
    db     = get_db()
    member = row_to_dict(db.execute("SELECT * FROM members WHERE id=?", (d["member_id"],)).fetchone())
    if not member:
        db.close(); return error("Member not found")
    if member["savings"] < amount:
        db.close(); return error(f"Insufficient balance. Current: KES {member['savings']:,.2f}")

    new_balance = member["savings"] - amount
    tid = gen_id("ST")
    ref = d.get("reference") or "WDR" + gen_id()

    db.execute("UPDATE members SET savings=? WHERE id=?", (new_balance, d["member_id"]))
    db.execute("UPDATE savings_accounts SET balance=? WHERE member_id=?", (new_balance, d["member_id"]))
    db.execute(
        "INSERT INTO savings_transactions (id,member_id,type,amount,category,txn_date,reference,balance_after,recorded_by) VALUES (?,?,?,?,?,?,?,?,?)",
        (tid, d["member_id"], "withdrawal", amount, d.get("category","voluntary"),
         d.get("date", date.today().isoformat()), ref, new_balance, g.user["sub"])
    )
    adjust_account_opening_balance(db, -amount)
    db.commit(); db.close()
    audit(f"Recorded withdrawal {tid}", "Savings", f"KES {amount} for {d['member_id']}")
    return success({"reference": ref, "new_balance": new_balance}, "Withdrawal processed", 201)

# ══════════════════════════════════════════════════════════════════════════════
# EXPENSES
# ══════════════════════════════════════════════════════════════════════════════